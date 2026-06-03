#!/usr/bin/env python3
"""
FLIM Phasor Analyzer — interactive phasor analysis & segmentation for CAM models
================================================================================

Designed for ex-ovo chicken CAM FLIM data (blood vessels / collagen / background),
but works for any single- or multi-channel TCSPC FLIM .ptu file.

Pipeline (phasorpy 0.10):
    .ptu -> pick channel -> TCSPC histogram -> phasor (g, s)
         -> reference calibration (phasorpy.phasor_calibrate)
         -> filter:  none | median (phasor_filter_median) | pawFLIM (phasor_filter_pawflim)
         -> min photon count filter (sum over H, absolute counts)
         -> phasor plot + intensity image
         -> segmentation: manual circular cursors  OR  Gaussian Mixture Model
         -> lifetimes per cluster: tau_phi, tau_mod, tau_normal (projected)
         -> annotate clusters + quantify (area %) + CSV / image export

Run:
    pip install phasorpy scikit-learn PySide6 matplotlib numpy pawflim openpyxl
    python flim_phasor_gui.py

Notes:
  * Channel selection: detected automatically from the PTU 'C' axis.
  * pawFLIM requires >=2 harmonics, so it computes the phasor at [H, 2H] on the
    calibrated, unfiltered coordinates (as the algorithm requires) and then uses
    the working harmonic H for analysis. It needs the optional `pawflim` package
    and a real loaded signal (not the synthetic demo).
  * Apparent lifetime, median filter and pawFLIM all use phasorpy's own
    implementations.

Author: Srijan
"""

import sys
import os
import csv
import time
import numpy as np

# ----------------------------------------------------------------------------
# Qt (PySide6) + matplotlib backend
# ----------------------------------------------------------------------------
try:
    from PySide6 import QtCore, QtGui, QtWidgets
    from PySide6.QtCore import Qt, Signal
except ImportError:
    sys.exit("PySide6 is required:  pip install PySide6")

os.environ.setdefault("QT_API", "pyside6")
import matplotlib
matplotlib.use("QtAgg")
from matplotlib.figure import Figure
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qtagg import NavigationToolbar2QT as NavigationToolbar
from matplotlib.patches import Circle, Ellipse

# ----------------------------------------------------------------------------
# phasorpy + sklearn
# ----------------------------------------------------------------------------
try:
    from phasorpy.io import signal_from_ptu
    from phasorpy.phasor import phasor_from_signal
    from phasorpy.lifetime import (
        phasor_calibrate,
        phasor_to_apparent_lifetime,
        phasor_from_lifetime,
        phasor_from_apparent_lifetime,
        phasor_to_normal_lifetime,
        phasor_to_lifetime_search,
    )
    from phasorpy.filter import (
        phasor_threshold,
        phasor_filter_median,
        phasor_filter_gaussian,
        phasor_filter_pawflim,
    )
    from phasorpy.cursor import mask_from_circular_cursor, pseudo_color
    from phasorpy.color import CATEGORICAL
except ImportError as e:
    sys.exit(f"phasorpy is required (pip install phasorpy). Import error: {e}")

try:
    from sklearn.mixture import GaussianMixture
    HAVE_SKLEARN = True
except ImportError:
    HAVE_SKLEARN = False


# ----------------------------------------------------------------------------
# PTU / reference caches (avoid re-decoding the same file on every Apply)
# ----------------------------------------------------------------------------
_REF_PTU_CACHE: dict[str, object] = {}          # normpath -> full integrated DataArray
_REF_PHASOR_CACHE: dict[tuple, tuple] = {}      # (normpath, channel, harmonics) -> (mean, real, imag)


def _ptu_cache_key(path: str) -> str:
    return os.path.normcase(os.path.abspath(path))


def load_reference_ptu(path: str):
    """Decode a reference .ptu once; reuse the in-memory histogram for later Apply calls."""
    key = _ptu_cache_key(path)
    if key not in _REF_PTU_CACHE:
        _REF_PTU_CACHE[key] = signal_from_ptu(path, channel=None, frame=-1, dtype=np.uint32)
    return _REF_PTU_CACHE[key]


def reference_phasor(ref_path: str, channel: int, harmonic):
    """Cached (mean, real, imag) for a reference file at the given channel and harmonic(s)."""
    harm_key = tuple(harmonic) if isinstance(harmonic, (list, tuple)) else (int(harmonic),)
    key = (_ptu_cache_key(ref_path), int(channel), harm_key)
    if key not in _REF_PHASOR_CACHE:
        rsig = reduce_signal(load_reference_ptu(ref_path), channel)
        rmean, rreal, rimag = phasor_from_signal(rsig, axis="H", harmonic=harmonic)
        _REF_PHASOR_CACHE[key] = (rmean, rreal, rimag)
    return _REF_PHASOR_CACHE[key]


def clear_ptu_caches():
    _REF_PTU_CACHE.clear()
    _REF_PHASOR_CACHE.clear()


# ----------------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------------
def to_2d(arr):
    """Squeeze length-1 dims; if still >2D, take leading index until 2D."""
    arr = np.squeeze(np.asarray(arr))
    while arr.ndim > 2:
        arr = arr[0]
    return arr


def categorical_rgb(i):
    c = np.asarray(CATEGORICAL)
    return tuple(float(x) for x in c[i % len(c)])


def reduce_signal(sig, channel):
    """Select a channel and collapse frames -> DataArray with dims (Y, X, H)."""
    if "C" in sig.dims:
        sig = sig.isel(C=int(min(channel, sig.sizes["C"] - 1)))
    if "T" in sig.dims:
        if sig.sizes["T"] > 1:
            sig = sig.sum("T")          # accumulate frames (more photons)
        else:
            sig = sig.squeeze("T", drop=True)
    # drop any other stray singleton dims except H
    for d in list(sig.dims):
        if d != "H" and sig.sizes.get(d, 2) == 1:
            sig = sig.squeeze(d, drop=True)
    return sig


def photon_count_from_signal(sig):
    """Total photon count per pixel = sum of the TCSPC histogram (axis H)."""
    if hasattr(sig, "sum") and "H" in getattr(sig, "dims", ()):
        pc = sig.sum(dim="H")
        for d in list(pc.dims):
            if d != "H" and pc.sizes.get(d, 1) == 1:
                pc = pc.squeeze(d, drop=True)
        return to_2d(np.asarray(pc.values, dtype=float))
    arr = np.asarray(sig, dtype=float)
    axis_h = -1
    return to_2d(np.sum(arr, axis=axis_h))


# ============================================================================
#  Data container
# ============================================================================
class PhasorData:
    def __init__(self):
        self.reset()

    def reset(self):
        self.signal_full = None       # raw multi-channel sample DataArray
        self.n_channels = 1
        self.channel = 0
        self.frequency = 80.0         # fundamental laser frequency (MHz)
        self.harmonic = 1
        self.is_synthetic = False
        # synthetic-only raw arrays (harmonic 1)
        self._syn = None              # (mean, real, imag)
        # processed/displayed arrays
        self.mean_raw = None       # per-pixel total photon count (sum over H); for display
        self.mean_thr = None       # photon count with NaN where below threshold (for reference)
        self.real_cal = None
        self.imag_cal = None
        # per-pixel apparent lifetime maps (ns), computed after thresholding
        self.tau_phi = None       # phase apparent lifetime
        self.tau_mod = None       # modulation apparent lifetime
        self.tau_normal = None    # normal lifetime (projection onto universal circle)
        self.sample_path = ""
        self.ref_path = ""
        self.ref_n_channels = 1
        self.ref_channel = 0

    # ---- loading -----------------------------------------------------------
    def load_sample(self, path):
        # channel=None keeps all channels for the dropdown; frame=-1 integrates
        # all frames for photon statistics; uint32 avoids overflow on integration
        sig = signal_from_ptu(path, channel=None, frame=-1, dtype=np.uint32)
        self.signal_full = sig
        self.is_synthetic = False
        self.sample_path = path
        self.n_channels = int(sig.sizes["C"]) if "C" in sig.dims else 1
        self.channel = 0
        self.frequency = float(sig.attrs.get("frequency", 80.0))
        # shape of one channel image
        red = reduce_signal(sig, 0)
        yx = [s for d, s in zip(red.dims, red.shape) if d != "H"]
        shape = tuple(yx) if len(yx) == 2 else (red.shape[0], red.shape[1])
        return shape, self.n_channels

    def load_synthetic(self, shape=(256, 256)):
        h, w = shape
        rng = np.random.default_rng(0)
        freq = 80.0
        self.frequency = freq
        self.harmonic = 1
        self.is_synthetic = True
        self.signal_full = None
        self.n_channels = 1
        self.channel = 0
        g = np.zeros(shape); s = np.zeros(shape); mean = np.zeros(shape)
        bg_g, bg_s = phasor_from_lifetime(freq, 0.4)
        g[:] = bg_g; s[:] = bg_s; mean[:] = rng.uniform(20, 80, shape)
        yy, xx = np.mgrid[0:h, 0:w]
        col_g, col_s = phasor_from_lifetime(freq, 0.25)
        for _ in range(6):
            cy, cx, r = rng.integers(0, h), rng.integers(0, w), rng.integers(15, 40)
            m = (yy - cy) ** 2 + (xx - cx) ** 2 < r ** 2
            g[m] = col_g; s[m] = col_s; mean[m] = rng.uniform(300, 900, m.sum())
        ves_g, ves_s = phasor_from_lifetime(freq, 2.0)
        for _ in range(5):
            y0 = rng.integers(0, h); x = np.arange(w)
            y = (y0 + 25 * np.sin(x / 30.0 + rng.uniform(0, 6))).astype(int) % h
            for dy in range(-3, 4):
                yi = (y + dy) % h
                g[yi, x] = ves_g; s[yi, x] = ves_s; mean[yi, x] = rng.uniform(400, 1000, w)
        g += rng.normal(0, 0.012, shape); s += rng.normal(0, 0.012, shape)
        self._syn = (mean, g, s)
        self.sample_path = "<synthetic demo>"
        return shape, 1

    # ---- channel signal helpers -------------------------------------------
    def _sample_channel_signal(self):
        return reduce_signal(self.signal_full, self.channel)

    def _reference_phasor(self, ref_path, harmonic):
        """Reference phasor coordinates (cached; does not re-read the .ptu each Apply)."""
        rmean, rreal, rimag = reference_phasor(ref_path, self.ref_channel, harmonic)
        if isinstance(harmonic, (list, tuple)):
            return rmean, rreal, rimag
        return to_2d(rmean), to_2d(rreal), to_2d(rimag)

    # ---- processing --------------------------------------------------------
    def apply_processing(self, ref_path=None, ref_lifetime=4.0, filter_mode="median",
                         median_size=3, median_repeat=1, paw_sigma=2.0, paw_levels=1,
                         intensity_min=0.0):
        H = int(self.harmonic)
        freq = float(self.frequency)

        if filter_mode == "pawflim" and not self.is_synthetic:
            # pawFLIM path: needs >=2 harmonics on calibrated, UNfiltered phasor
            harmonics = [H, 2 * H]
            sig = self._sample_channel_signal()
            photon_count = photon_count_from_signal(sig)
            mean, real, imag = phasor_from_signal(sig, axis="H", harmonic=harmonics)
            if ref_path:
                rmean, rreal, rimag = self._reference_phasor(ref_path, harmonics)
                real, imag = phasor_calibrate(
                    real, imag, rmean, rreal, rimag,
                    frequency=freq, lifetime=float(ref_lifetime), harmonic=harmonics)
                self.ref_path = ref_path
            mean, real, imag = phasor_filter_pawflim(
                mean, real, imag, sigma=float(paw_sigma),
                levels=int(paw_levels), harmonic=harmonics)
            # use working harmonic (index 0 == harmonic H)
            real = to_2d(real[0]); imag = to_2d(imag[0]); mean = to_2d(mean)
        else:
            # single-harmonic path: none / median
            if self.is_synthetic:
                mean, real, imag = (a.astype(float).copy() for a in self._syn)
                mean, real, imag = to_2d(mean), to_2d(real), to_2d(imag)
                photon_count = mean.copy()
            else:
                sig = self._sample_channel_signal()
                photon_count = photon_count_from_signal(sig)
                mean, real, imag = phasor_from_signal(sig, axis="H", harmonic=H)
                mean, real, imag = to_2d(mean), to_2d(real), to_2d(imag)
            if ref_path and not self.is_synthetic:
                rmean, rreal, rimag = self._reference_phasor(ref_path, H)
                real, imag = phasor_calibrate(
                    real, imag, rmean, rreal, rimag,
                    frequency=freq, lifetime=float(ref_lifetime), harmonic=H)
                self.ref_path = ref_path
            if filter_mode == "median" and median_size >= 1 and median_repeat >= 1:
                mean, real, imag = phasor_filter_median(
                    mean, real, imag, size=int(median_size), repeat=int(median_repeat))
            elif filter_mode == "gaussian" and median_size >= 1 and median_repeat >= 1:
                mean, real, imag = phasor_filter_gaussian(
                    mean, real, imag, size=int(median_size), repeat=int(median_repeat))

        # Photon filter uses total counts (sum over H), not phasor DC (which is ~counts/n_bins)
        real = to_2d(real)
        imag = to_2d(imag)
        photon_count = to_2d(photon_count)
        self.mean_raw = np.asarray(photon_count, dtype=float)

        thr = float(intensity_min)
        if thr > 0:
            mean_thr, real, imag = phasor_threshold(
                photon_count, real, imag, mean_min=thr, detect_harmonics=False)
        else:
            mean_thr = self.mean_raw.copy()
            bad = ~np.isfinite(real) | ~np.isfinite(imag)
            if np.any(bad):
                mean_thr = mean_thr.copy()
                mean_thr[bad] = np.nan
                real = np.where(bad, np.nan, real)
                imag = np.where(bad, np.nan, imag)

        self.mean_thr = to_2d(mean_thr)
        self.real_cal = to_2d(real)
        self.imag_cal = to_2d(imag)
        self._intensity_stats = self._compute_intensity_stats(thr)

        # per-pixel apparent / normal lifetime maps at the working frequency
        work_freq = freq * H
        with np.errstate(invalid="ignore", divide="ignore"):
            tau_phi, tau_mod = phasor_to_apparent_lifetime(real, imag, work_freq)
            tau_normal = phasor_to_normal_lifetime(real, imag, work_freq)
        self.tau_phi = np.asarray(tau_phi, dtype=float)
        self.tau_mod = np.asarray(tau_mod, dtype=float)
        self.tau_normal = np.asarray(tau_normal, dtype=float)

    @property
    def work_frequency(self):
        return self.frequency * self.harmonic

    def valid_mask(self):
        return np.isfinite(self.real_cal) & np.isfinite(self.imag_cal)

    def _compute_intensity_stats(self, threshold):
        """Photon-count range and fraction masked by the intensity threshold."""
        if self.mean_raw is None:
            return {}
        raw = np.asarray(self.mean_raw, dtype=float)
        finite = raw[np.isfinite(raw)]
        if finite.size == 0:
            return {"threshold": threshold, "masked_pct": 100.0}
        n_pixels = int(finite.size)
        n_below = int(np.sum(finite < threshold)) if threshold > 0 else 0
        masked_pct = 100.0 * n_below / n_pixels if n_pixels else 0.0
        return {
            "threshold": threshold,
            "min": float(np.min(finite)),
            "median": float(np.median(finite)),
            "max": float(np.max(finite)),
            "n_pixels": n_pixels,
            "n_below": n_below,
            "masked_pct": masked_pct,
        }


# ============================================================================
#  Phasor plot canvas with interactive circular cursors
# ============================================================================
class PhasorCanvas(FigureCanvas):
    cursorChanged = Signal()   # committed: gesture end / add / remove (full recompute)
    cursorMoving = Signal()    # interactive: drag / scroll / slider (fast overlay only)

    def __init__(self, parent=None):
        self.fig = Figure(figsize=(5, 5), tight_layout=True)
        super().__init__(self.fig)
        self.setParent(parent)
        self.ax = self.fig.add_subplot(111)
        self.data = None
        self.cursors = []
        self.selected = -1
        self._dragging = False
        self._gmm_artists = []
        self._init_axes()
        self.mpl_connect("button_press_event", self.on_press)
        self.mpl_connect("button_release_event", self.on_release)
        self.mpl_connect("motion_notify_event", self.on_motion)
        self.mpl_connect("scroll_event", self.on_scroll)

    def _init_axes(self):
        self.ax.clear()
        self.ax.set_xlim(-0.05, 1.05)
        self.ax.set_ylim(-0.05, 0.75)
        self.ax.set_xlabel("g"); self.ax.set_ylabel("s")
        self.ax.set_aspect("equal"); self.ax.set_title("Phasor")
        theta = np.linspace(0, np.pi, 200)
        self.ax.plot(0.5 + 0.5 * np.cos(theta), 0.5 * np.sin(theta), "k-", lw=1, alpha=0.6)
        self.ax.grid(alpha=0.2)

    def set_data(self, data):
        self.data = data
        self.redraw_hist()

    def redraw_hist(self):
        self._init_axes()
        if self.data is not None and self.data.real_cal is not None:
            g, s = self.data.real_cal, self.data.imag_cal
            m = self.data.valid_mask()
            if m.sum() > 0:
                self.ax.hist2d(g[m].ravel(), s[m].ravel(), bins=256,
                               range=[[-0.05, 1.05], [-0.05, 0.75]], cmap="turbo", cmin=1)
            if freq_ok(self.data):
                for tau in (0.5, 1, 2, 3, 4, 8):
                    gg, ss = phasor_from_lifetime(self.data.work_frequency, tau)
                    self.ax.plot(gg, ss, "k.", ms=4)
                    self.ax.annotate(f"{tau}ns", (gg, ss), fontsize=7, alpha=0.7)
        self._redraw_cursors()
        self.draw_idle()

    def add_cursor(self, radius=0.05):
        if self.data is not None and self.data.valid_mask().sum() > 0:
            m = self.data.valid_mask()
            cr = float(np.nanmedian(self.data.real_cal[m]))
            ci = float(np.nanmedian(self.data.imag_cal[m]))
        else:
            cr, ci = 0.5, 0.3
        idx = len(self.cursors)
        self.cursors.append(dict(center_real=cr, center_imag=ci, radius=radius,
                                 color=categorical_rgb(idx),
                                 label=f"cluster {idx + 1}", patch=None, label_artist=None))
        self.selected = idx
        self._redraw_cursors(); self.draw_idle(); self.cursorChanged.emit()

    def remove_selected(self):
        if 0 <= self.selected < len(self.cursors):
            self._remove_cursor_artists(self.cursors[self.selected])
            self.cursors.pop(self.selected)
            self.selected = len(self.cursors) - 1 if self.cursors else -1
            self._redraw_cursors(); self.draw(); self.cursorChanged.emit()

    def clear_cursors(self):
        for c in self.cursors:
            self._remove_cursor_artists(c)
        self.cursors = []
        self.selected = -1
        self._purge_stray_cursor_artists()
        self.draw(); self.cursorChanged.emit()

    def set_selected_radius(self, r):
        if 0 <= self.selected < len(self.cursors):
            self.cursors[self.selected]["radius"] = float(r)
            self._redraw_cursors(); self.draw_idle(); self.cursorMoving.emit()

    def _remove_cursor_artists(self, c):
        for key in ("patch", "label_artist"):
            art = c.get(key)
            if art is not None:
                try:
                    art.remove()
                except Exception:
                    pass
                c[key] = None

    def _purge_stray_cursor_artists(self):
        """Remove circle/label artists left behind after a cursor was deleted."""
        keep_patches = {c["patch"] for c in self.cursors if c.get("patch")}
        keep_labels = {c["label_artist"] for c in self.cursors if c.get("label_artist")}
        for p in list(self.ax.patches):
            if isinstance(p, Circle) and p not in keep_patches:
                try:
                    p.remove()
                except Exception:
                    pass
        for t in list(self.ax.texts):
            if t not in keep_labels:
                try:
                    t.remove()
                except Exception:
                    pass

    def _redraw_cursors(self):
        for c in self.cursors:
            self._remove_cursor_artists(c)
        self._purge_stray_cursor_artists()
        for i, c in enumerate(self.cursors):
            lw = 2.5 if i == self.selected else 1.2
            patch = Circle((c["center_real"], c["center_imag"]), c["radius"],
                           fill=False, edgecolor=c["color"], lw=lw)
            self.ax.add_patch(patch)
            c["patch"] = patch
            ann = self.ax.annotate(
                str(i + 1), (c["center_real"], c["center_imag"]),
                color=c["color"], fontsize=9, ha="center", va="center", weight="bold")
            c["label_artist"] = ann

    def show_gmm(self, means, covs, colors):
        self.clear_gmm()
        for k in range(len(means)):
            mg, ms = means[k]; cov = covs[k]
            vals, vecs = np.linalg.eigh(cov)
            order = vals.argsort()[::-1]; vals, vecs = vals[order], vecs[:, order]
            angle = np.degrees(np.arctan2(vecs[1, 0], vecs[0, 0]))
            for n in (1, 2):
                w, h = 2 * n * np.sqrt(np.maximum(vals, 1e-9))
                e = Ellipse((mg, ms), w, h, angle=angle, fill=False,
                            edgecolor=colors[k], lw=1.5, alpha=0.9 / n)
                self.ax.add_patch(e); self._gmm_artists.append(e)
            pt = self.ax.plot(mg, ms, "x", color=colors[k], ms=8, mew=2)[0]
            self._gmm_artists.append(pt)
        self.draw_idle()

    def clear_gmm(self):
        for a in self._gmm_artists:
            try: a.remove()
            except Exception: pass
        self._gmm_artists = []
        self.draw_idle()

    def _hit(self, x, y):
        for i, c in enumerate(self.cursors):
            if np.hypot(x - c["center_real"], y - c["center_imag"]) <= c["radius"] * 1.3:
                return i
        return -1

    def on_press(self, event):
        if event.inaxes != self.ax or event.xdata is None:
            return
        i = self._hit(event.xdata, event.ydata)
        if event.button == 3 and i >= 0:
            self._remove_cursor_artists(self.cursors[i])
            self.cursors.pop(i)
            self.selected = len(self.cursors) - 1 if self.cursors else -1
            self._redraw_cursors(); self.draw(); self.cursorChanged.emit()
            return
        if event.button != 1:
            return
        if i >= 0:
            self.selected = i
            self._dragging = True
            self._redraw_cursors(); self.draw_idle(); self.cursorChanged.emit()
        else:
            self.selected = -1
            self._redraw_cursors(); self.draw_idle()

    def on_motion(self, event):
        if not self._dragging or event.inaxes != self.ax or event.xdata is None:
            return
        if 0 <= self.selected < len(self.cursors):
            self.cursors[self.selected]["center_real"] = float(event.xdata)
            self.cursors[self.selected]["center_imag"] = float(event.ydata)
            self._redraw_cursors(); self.draw_idle(); self.cursorMoving.emit()

    def on_release(self, event):
        if self._dragging:
            self._dragging = False; self.cursorChanged.emit()

    def on_scroll(self, event):
        if event.inaxes != self.ax or not (0 <= self.selected < len(self.cursors)):
            return
        step = 0.005 if event.button == "up" else -0.005
        r = max(0.005, self.cursors[self.selected]["radius"] + step)
        self.cursors[self.selected]["radius"] = r
        self._redraw_cursors(); self.draw_idle(); self.cursorMoving.emit()


def freq_ok(data):
    return data is not None and data.frequency and data.work_frequency > 0


# ============================================================================
#  Image canvas
# ============================================================================
class ImageCanvas(FigureCanvas):
    def __init__(self, parent=None):
        self.fig = Figure(figsize=(5, 5), tight_layout=True)
        super().__init__(self.fig)
        self.setParent(parent)
        self.ax = self.fig.add_subplot(111)
        self.ax.set_title("Image"); self.ax.axis("off")
        self._im = None
        self._cbar = None

    def _reset_ax(self, title):
        """Rebuild the axes from scratch so any previous colorbar is dropped."""
        self.fig.clear()
        self.ax = self.fig.add_subplot(111)
        self.ax.set_title(title); self.ax.axis("off")
        self._cbar = None

    def show_intensity(self, mean, vmin=None, vmax=None):
        """Show photon image; NaN = masked (shown dark), scaling from finite pixels only."""
        self._reset_ax("Intensity")
        arr = np.squeeze(np.asarray(mean, dtype=float))
        while arr.ndim > 2:
            arr = arr[0]
        finite = arr[np.isfinite(arr)]
        if finite.size == 0:
            arr = np.zeros((2, 2))
            finite = arr.ravel()
        if vmin is None or vmax is None:
            lo, hi = np.percentile(finite, [2, 98])
            vmin = lo if vmin is None else vmin
            vmax = hi if vmax is None else vmax
        if vmax <= vmin:
            vmax = vmin + 1.0
        masked = np.ma.masked_invalid(arr)
        self._im = self.ax.imshow(masked, cmap="gray", vmin=vmin, vmax=vmax)
        self.draw_idle()

    def show_map(self, arr, title, cmap="viridis", label="ns", vmin=None, vmax=None):
        """Display a per-pixel scalar map (e.g. apparent lifetime) with a colorbar."""
        self._reset_ax(title)
        self._im = self.ax.imshow(arr, cmap=cmap, vmin=vmin, vmax=vmax)
        self._cbar = self.fig.colorbar(self._im, ax=self.ax, fraction=0.046, pad=0.04)
        self._cbar.set_label(label)
        self.draw_idle()

    def show_overlay(self, overlay, title="Segmentation"):
        self._reset_ax(title)
        self._im = self.ax.imshow(overlay)
        self.draw_idle()

    def update_overlay(self, overlay, title="Segmentation"):
        """Fast path for live updates: reuse the AxesImage via set_data."""
        arr = None if self._im is None else self._im.get_array()
        if (self._cbar is None and self._im is not None and arr is not None
                and tuple(arr.shape) == tuple(overlay.shape)):
            self._im.set_data(overlay)
        else:
            self._reset_ax(title)
            self._im = self.ax.imshow(overlay)
        self.draw_idle()


# ============================================================================
#  Main window
# ============================================================================
class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("FLIM Phasor Analyzer — CAM segmentation")
        self.resize(1500, 980)
        self.data = PhasorData()
        self.datasets = []        # multi-image mode: list of PhasorData
        self.active_idx = -1
        self.last_overlay = None
        self.cluster_stats = []
        self.mode = "cursor"
        self._label_signal_connected = False
        self._paint_timer = QtCore.QTimer(self)
        self._paint_timer.setSingleShot(True)
        self._paint_timer.setInterval(250)   # ms after interaction stops -> full recompute
        self._paint_timer.timeout.connect(self._deferred_full_compute)
        self._build_ui()
        QtGui.QShortcut(QtGui.QKeySequence(Qt.Key.Key_Delete), self, self.remove_cursor)
        QtGui.QShortcut(QtGui.QKeySequence(Qt.Key.Key_Backspace), self, self.remove_cursor)

    # ---- UI ----------------------------------------------------------------
    def _build_ui(self):
        central = QtWidgets.QWidget(); self.setCentralWidget(central)
        main = QtWidgets.QHBoxLayout(central)

        panel = QtWidgets.QWidget(); panel.setFixedWidth(340)
        pl = QtWidgets.QVBoxLayout(panel); pl.setAlignment(Qt.AlignmentFlag.AlignTop)

        # ---- files ----
        gb_files = QtWidgets.QGroupBox("1 · Files")
        fl = QtWidgets.QGridLayout(gb_files)
        self.lbl_sample = QtWidgets.QLabel("(no sample)"); self.lbl_sample.setWordWrap(True)
        self.lbl_ref = QtWidgets.QLabel("(no reference)"); self.lbl_ref.setWordWrap(True)
        btn_sample = QtWidgets.QPushButton("Choose sample .ptu…"); btn_sample.clicked.connect(self.choose_sample)
        btn_ref = QtWidgets.QPushButton("Choose reference .ptu…"); btn_ref.clicked.connect(self.choose_ref)
        btn_demo = QtWidgets.QPushButton("Load synthetic demo"); btn_demo.clicked.connect(self.load_demo)
        fl.addWidget(btn_sample, 0, 0, 1, 2); fl.addWidget(self.lbl_sample, 1, 0, 1, 2)
        fl.addWidget(QtWidgets.QLabel("Sample channel"), 2, 0)
        self.cb_channel = QtWidgets.QComboBox(); self.cb_channel.addItem("0")
        self.cb_channel.currentIndexChanged.connect(self.on_channel_change)
        fl.addWidget(self.cb_channel, 2, 1)
        fl.addWidget(btn_ref, 3, 0, 1, 2); fl.addWidget(self.lbl_ref, 4, 0, 1, 2)
        fl.addWidget(QtWidgets.QLabel("Ref channel"), 5, 0)
        self.cb_ref_channel = QtWidgets.QComboBox(); self.cb_ref_channel.addItem("0")
        self.cb_ref_channel.setEnabled(False)
        self.cb_ref_channel.currentIndexChanged.connect(self.on_ref_channel_change)
        fl.addWidget(self.cb_ref_channel, 5, 1)
        fl.addWidget(btn_demo, 6, 0, 1, 2)
        self.chk_multi = QtWidgets.QCheckBox("Multi-image mode (keep & toggle several)")
        self.chk_multi.toggled.connect(self.on_multi_toggle)
        fl.addWidget(self.chk_multi, 7, 0, 1, 2)
        self.multi_box = QtWidgets.QWidget()
        mbl = QtWidgets.QHBoxLayout(self.multi_box); mbl.setContentsMargins(0, 0, 0, 0)
        self.cb_image = QtWidgets.QComboBox()
        self.cb_image.currentIndexChanged.connect(self.on_image_combo_change)
        btn_rmimg = QtWidgets.QPushButton("Remove"); btn_rmimg.clicked.connect(self.remove_image)
        mbl.addWidget(self.cb_image, 1); mbl.addWidget(btn_rmimg)
        self.multi_box.setVisible(False)
        fl.addWidget(self.multi_box, 8, 0, 1, 2)
        pl.addWidget(gb_files)

        # ---- calibration / preprocessing ----
        gb_proc = QtWidgets.QGroupBox("2 · Calibration & preprocessing")
        prl = QtWidgets.QFormLayout(gb_proc)
        self.sp_harm = QtWidgets.QSpinBox(); self.sp_harm.setRange(1, 8); self.sp_harm.setValue(1)
        self.sp_freq = QtWidgets.QDoubleSpinBox(); self.sp_freq.setRange(1, 1000); self.sp_freq.setDecimals(3); self.sp_freq.setValue(80.0); self.sp_freq.setSuffix(" MHz")
        self.sp_reflt = QtWidgets.QDoubleSpinBox(); self.sp_reflt.setRange(0.0, 100.0); self.sp_reflt.setDecimals(3); self.sp_reflt.setValue(4.0); self.sp_reflt.setSuffix(" ns")
        self.sp_reflt.setToolTip(
            "Known lifetime of the reference dye (ns), e.g. ~4 for fluorescein. "
            "Wrong value shifts the phasor cloud vertically on the plot.")
        self.cb_filter = QtWidgets.QComboBox(); self.cb_filter.addItems(["none", "median", "gaussian", "pawflim"])
        self.cb_filter.setCurrentText("median"); self.cb_filter.currentTextChanged.connect(self.on_filter_change)
        self.sp_msize = QtWidgets.QSpinBox(); self.sp_msize.setRange(3, 11); self.sp_msize.setSingleStep(2); self.sp_msize.setValue(3)
        self.sp_mrep = QtWidgets.QSpinBox(); self.sp_mrep.setRange(1, 10); self.sp_mrep.setValue(1)
        self.sp_psigma = QtWidgets.QDoubleSpinBox(); self.sp_psigma.setRange(0.5, 6.0); self.sp_psigma.setSingleStep(0.5); self.sp_psigma.setValue(2.0)
        self.sp_plevels = QtWidgets.QSpinBox(); self.sp_plevels.setRange(1, 6); self.sp_plevels.setValue(1)
        self.sp_thr = QtWidgets.QSpinBox()
        self.sp_thr.setRange(0, 2_000_000_000)
        self.sp_thr.setSingleStep(100)
        self.sp_thr.setValue(0)
        self.sp_thr.setToolTip(
            "Minimum total photon count per pixel (sum of the TCSPC histogram). "
            "Pixels below this count are excluded from the phasor plot and segmentation only; "
            "the intensity image always shows all photons. 0 = off.")
        self.lbl_photon_range = QtWidgets.QLabel("(apply to see photon range)")
        self.lbl_photon_range.setWordWrap(True)
        self.lbl_photon_range.setStyleSheet("color: gray; font-size: 11px;")
        prl.addRow("Harmonic", self.sp_harm)
        prl.addRow("Laser freq", self.sp_freq)
        prl.addRow("Ref lifetime", self.sp_reflt)
        prl.addRow("Filter", self.cb_filter)
        self.row_msize = self._form_row(prl, "Kernel size", self.sp_msize)
        self.row_mrep = self._form_row(prl, "Repeat", self.sp_mrep)
        self.row_psigma = self._form_row(prl, "pawFLIM sigma", self.sp_psigma)
        self.row_plevels = self._form_row(prl, "pawFLIM levels", self.sp_plevels)
        prl.addRow("Min photon count", self.sp_thr)
        prl.addRow("", self.lbl_photon_range)
        btn_apply = QtWidgets.QPushButton("Apply / recompute phasor"); btn_apply.clicked.connect(self.apply_processing)
        prl.addRow(btn_apply)
        pl.addWidget(gb_proc)
        self.on_filter_change("median")

        # ---- mode ----
        gb_mode = QtWidgets.QGroupBox("3 · Segmentation mode")
        ml = QtWidgets.QVBoxLayout(gb_mode)
        self.rb_cursor = QtWidgets.QRadioButton("Manual circular cursors")
        self.rb_gmm = QtWidgets.QRadioButton("Gaussian Mixture Model")
        self.rb_cursor.setChecked(True); self.rb_cursor.toggled.connect(self.on_mode_change)
        ml.addWidget(self.rb_cursor); ml.addWidget(self.rb_gmm)

        self.cursor_box = QtWidgets.QWidget()
        cbl = QtWidgets.QGridLayout(self.cursor_box); cbl.setContentsMargins(0, 0, 0, 0)
        b_add = QtWidgets.QPushButton("+ Add circle"); b_add.clicked.connect(self.add_cursor)
        b_del = QtWidgets.QPushButton("Delete selected"); b_del.clicked.connect(self.remove_cursor)
        b_del.setToolTip("Select a circle (click it), then press Delete or this button.")
        b_clr = QtWidgets.QPushButton("Clear all circles"); b_clr.clicked.connect(self.clear_cursors)
        cbl.addWidget(b_add, 0, 0); cbl.addWidget(b_del, 0, 1); cbl.addWidget(b_clr, 0, 2)
        hint = QtWidgets.QLabel("Click a circle to select · drag to move · "
                                "scroll to resize · right-click or Del to delete")
        hint.setWordWrap(True); hint.setStyleSheet("color: gray; font-size: 11px;")
        cbl.addWidget(hint, 1, 0, 1, 3)
        cbl.addWidget(QtWidgets.QLabel("Selected radius"), 2, 0, 1, 3)
        self.sld_radius = QtWidgets.QSlider(Qt.Orientation.Horizontal); self.sld_radius.setRange(5, 400); self.sld_radius.setValue(50)
        self.sld_radius.valueChanged.connect(self.on_radius_slider)
        self.lbl_radius = QtWidgets.QLabel("0.050")
        cbl.addWidget(self.sld_radius, 3, 0, 1, 2); cbl.addWidget(self.lbl_radius, 3, 2)
        ml.addWidget(self.cursor_box)

        self.gmm_box = QtWidgets.QWidget()
        gbl = QtWidgets.QGridLayout(self.gmm_box); gbl.setContentsMargins(0, 0, 0, 0)
        gbl.addWidget(QtWidgets.QLabel("Components"), 0, 0)
        self.sp_ncomp = QtWidgets.QSpinBox(); self.sp_ncomp.setRange(1, 12); self.sp_ncomp.setValue(3)
        gbl.addWidget(self.sp_ncomp, 0, 1)
        gbl.addWidget(QtWidgets.QLabel("Covariance"), 1, 0)
        self.cb_cov = QtWidgets.QComboBox(); self.cb_cov.addItems(["full", "tied", "diag", "spherical"])
        gbl.addWidget(self.cb_cov, 1, 1)
        self.chk_bic = QtWidgets.QCheckBox("auto-select n by BIC (1..N)"); gbl.addWidget(self.chk_bic, 2, 0, 1, 2)
        b_fit = QtWidgets.QPushButton("Fit GMM + paint"); b_fit.clicked.connect(self.fit_gmm)
        b_clr_gmm = QtWidgets.QPushButton("Clear GMM"); b_clr_gmm.clicked.connect(self.clear_gmm)
        gbl.addWidget(b_fit, 3, 0, 1, 2); gbl.addWidget(b_clr_gmm, 4, 0, 1, 2)
        gmm_hint = QtWidgets.QLabel("Switch to this mode, set components, then Fit. "
                                    "Clears manual circles.")
        gmm_hint.setWordWrap(True); gmm_hint.setStyleSheet("color: gray; font-size: 11px;")
        gbl.addWidget(gmm_hint, 5, 0, 1, 2)
        self.gmm_box.setVisible(False); ml.addWidget(self.gmm_box)
        pl.addWidget(gb_mode)

        # ---- actions ----
        gb_act = QtWidgets.QGroupBox("4 · Compute & export")
        al = QtWidgets.QVBoxLayout(gb_act)
        b_paint = QtWidgets.QPushButton("Paint + compute lifetimes"); b_paint.clicked.connect(self.compute_and_paint)
        al.addWidget(b_paint)
        self.chk_live = QtWidgets.QCheckBox("live paint while moving circles"); self.chk_live.setChecked(True)
        al.addWidget(self.chk_live)
        self.chk_overlay = QtWidgets.QCheckBox("show overlay (else base image)"); self.chk_overlay.setChecked(True)
        self.chk_overlay.stateChanged.connect(self.refresh_image); al.addWidget(self.chk_overlay)
        base_row = QtWidgets.QHBoxLayout()
        base_row.addWidget(QtWidgets.QLabel("Base image"))
        self.cb_base = QtWidgets.QComboBox()
        self.cb_base.addItems(["Intensity", "Apparent \u03c4\u03c6 (ns)",
                               "Apparent \u03c4m (ns)", "Normal \u03c4 (ns)"])
        self.cb_base.currentIndexChanged.connect(self.refresh_image)
        base_row.addWidget(self.cb_base, 1); al.addLayout(base_row)
        row = QtWidgets.QHBoxLayout()
        b1 = QtWidgets.QPushButton("Save overlay"); b1.clicked.connect(self.save_overlay)
        b2 = QtWidgets.QPushButton("Save phasor"); b2.clicked.connect(self.save_phasor)
        row.addWidget(b1); row.addWidget(b2); al.addLayout(row)
        rowx = QtWidgets.QHBoxLayout()
        b3 = QtWidgets.QPushButton("Export CSV"); b3.clicked.connect(self.export_csv)
        b4 = QtWidgets.QPushButton("Export Excel"); b4.clicked.connect(self.export_xlsx)
        rowx.addWidget(b3); rowx.addWidget(b4); al.addLayout(rowx)
        pl.addWidget(gb_act); pl.addStretch(1)
        main.addWidget(panel)

        # ---- plots ----
        plots = QtWidgets.QSplitter(Qt.Orientation.Horizontal)
        lw = QtWidgets.QWidget(); lv = QtWidgets.QVBoxLayout(lw)
        self.phasor = PhasorCanvas(self)
        self.phasor.cursorChanged.connect(self.on_cursor_changed)
        self.phasor.cursorMoving.connect(self.on_cursor_moving)
        self.phasor.mpl_connect("key_press_event", self._on_phasor_key_press)
        self.phasor.setFocusPolicy(Qt.FocusPolicy.StrongFocus)
        lv.addWidget(NavigationToolbar(self.phasor, self)); lv.addWidget(self.phasor)
        rw = QtWidgets.QWidget(); rv = QtWidgets.QVBoxLayout(rw)
        self.image = ImageCanvas(self)
        rv.addWidget(NavigationToolbar(self.image, self)); rv.addWidget(self.image)
        plots.addWidget(lw); plots.addWidget(rw); plots.setSizes([700, 700])

        self.table = QtWidgets.QTableWidget(0, 10)
        self.table.setHorizontalHeaderLabels(
            ["#", "Color", "Label (what you see)", "g", "s",
             "tau_phi (ns)", "tau_mod (ns)", "tau_normal (ns)",
             "Pixels", "Area %"])
        self.table.horizontalHeader().setSectionResizeMode(2, QtWidgets.QHeaderView.ResizeMode.Stretch)
        self.table.setMaximumHeight(220)

        rightside = QtWidgets.QSplitter(Qt.Orientation.Vertical)
        rightside.addWidget(plots); rightside.addWidget(self.table); rightside.setSizes([720, 220])
        main.addWidget(rightside, 1)

        self.status = self.statusBar()
        self.status.showMessage("Ready. Load a .ptu sample (and reference) or try the synthetic demo.")

    def _form_row(self, form, label, widget):
        """Add a labelled row and return a (label_widget, field_widget) tuple for show/hide."""
        lbl = QtWidgets.QLabel(label)
        form.addRow(lbl, widget)
        return (lbl, widget)

    def _run_busy(self, message: str, fn):
        """Run a blocking call with a modal progress dialog; return (result, seconds)."""
        dlg = QtWidgets.QProgressDialog(message, None, 0, 0, self)
        dlg.setWindowModality(Qt.WindowModality.WindowModal)
        dlg.setMinimumDuration(0)
        dlg.setCancelButton(None)
        dlg.show()
        QtWidgets.QApplication.processEvents()
        t0 = time.perf_counter()
        try:
            return fn(), time.perf_counter() - t0
        finally:
            dlg.close()

    @staticmethod
    def _fmt_elapsed(seconds: float) -> str:
        if seconds < 1.0:
            return f"{seconds * 1000:.0f} ms"
        return f"{seconds:.2f} s"

    # ---- show/hide filter params ------------------------------------------
    def on_filter_change(self, mode):
        is_kernel = mode in ("median", "gaussian"); is_paw = mode == "pawflim"
        for lbl, w in (self.row_msize, self.row_mrep):
            lbl.setVisible(is_kernel); w.setVisible(is_kernel)
        for lbl, w in (self.row_psigma, self.row_plevels):
            lbl.setVisible(is_paw); w.setVisible(is_paw)

    # ---- mode switching ----------------------------------------------------
    def on_mode_change(self):
        self.mode = "cursor" if self.rb_cursor.isChecked() else "gmm"
        self.cursor_box.setVisible(self.mode == "cursor")
        self.gmm_box.setVisible(self.mode == "gmm")
        if self.mode == "cursor":
            self.clear_gmm()
        else:
            self.phasor.clear_cursors()
            self.last_overlay = None
            self.cluster_stats = []
            self._fill_table()

    # ---- file actions ------------------------------------------------------
    def choose_sample(self):
        path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, "Choose sample FLIM file", "", "PicoQuant PTU (*.ptu);;All files (*)")
        if not path:
            return
        d = PhasorData()
        try:
            (shape, nch), t_load = self._run_busy(
                f"Decoding sample .ptu ({os.path.basename(path)})…",
                lambda: d.load_sample(path))
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Load error", str(e)); return
        self._activate_new_dataset(d)
        _, t_proc = self._run_busy("Computing phasor…", self.apply_processing)
        self.status.showMessage(
            f"Loaded {os.path.basename(path)} — {shape[1]}×{shape[0]}, {nch} ch, "
            f"{d.frequency:.2f} MHz  "
            f"(decode {self._fmt_elapsed(t_load)}, phasor {self._fmt_elapsed(t_proc)})")

    def choose_ref(self):
        path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, "Choose reference FLIM file", "", "PicoQuant PTU (*.ptu);;All files (*)")
        if not path:
            return
        try:
            rsig, t_ref = self._run_busy(
                f"Decoding reference .ptu ({os.path.basename(path)})…",
                lambda: load_reference_ptu(path))
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Reference load error", str(e)); return
        ref_nch = int(rsig.sizes["C"]) if "C" in rsig.dims else 1
        self.data.ref_path = path
        self.data.ref_n_channels = ref_nch
        self.data.ref_channel = min(self.data.channel, ref_nch - 1)
        self.lbl_ref.setText(os.path.basename(path))
        self._update_ref_channel_combo()
        if self.data.signal_full is not None or self.data.is_synthetic:
            self.apply_processing()
            self.status.showMessage(
                f"Reference ch {self.data.ref_channel} cached ({self._fmt_elapsed(t_ref)}); "
                "phasor recalibrated.")
        else:
            self.status.showMessage(
                f"Reference cached — {ref_nch} channel(s) ({self._fmt_elapsed(t_ref)}). "
                "Load a sample, then Apply.")

    def load_demo(self):
        d = PhasorData()
        d.load_synthetic()
        self._activate_new_dataset(d)
        self.lbl_ref.setText("(none — already calibrated)")
        if self.cb_filter.currentText() == "pawflim":
            self.cb_filter.setCurrentText("median")
        self.apply_processing()
        self.status.showMessage("Synthetic CAM-like demo loaded "
                                "(background 0.4 ns / collagen 0.25 ns / vessels 2.0 ns).")

    # ---- multi-image management -------------------------------------------
    def _activate_new_dataset(self, d):
        """Make d the active dataset; append to the set if multi-image mode is on."""
        if self.chk_multi.isChecked():
            self.datasets.append(d)
            self.active_idx = len(self.datasets) - 1
        self.data = d
        self._restore_ui_for_active()
        if self.chk_multi.isChecked():
            self._refresh_image_combo()

    def _restore_ui_for_active(self):
        d = self.data
        self.lbl_sample.setText(os.path.basename(d.sample_path) or "<synthetic demo>")
        self.lbl_ref.setText(os.path.basename(d.ref_path) if d.ref_path else "(none)")
        nch = max(1, d.n_channels)
        self.cb_channel.blockSignals(True)
        self.cb_channel.clear()
        self.cb_channel.addItems([str(i) for i in range(nch)])
        self.cb_channel.setCurrentIndex(min(d.channel, nch - 1))
        self.cb_channel.blockSignals(False)
        self._update_ref_channel_combo()
        self.sp_freq.setValue(d.frequency)
        self.sp_harm.setValue(d.harmonic)

    def _update_ref_channel_combo(self):
        d = self.data
        has_ref = bool(d.ref_path)
        nch = max(1, d.ref_n_channels) if has_ref else 1
        self.cb_ref_channel.blockSignals(True)
        self.cb_ref_channel.clear()
        self.cb_ref_channel.addItems([str(i) for i in range(nch)])
        if has_ref:
            self.cb_ref_channel.setCurrentIndex(min(d.ref_channel, nch - 1))
        self.cb_ref_channel.setEnabled(has_ref)
        self.cb_ref_channel.blockSignals(False)

    def _refresh_image_combo(self):
        self.cb_image.blockSignals(True)
        self.cb_image.clear()
        for i, d in enumerate(self.datasets):
            name = os.path.basename(d.sample_path) or f"image {i + 1}"
            self.cb_image.addItem(f"{i + 1}: {name}")
        if 0 <= self.active_idx < len(self.datasets):
            self.cb_image.setCurrentIndex(self.active_idx)
        self.cb_image.blockSignals(False)

    def on_multi_toggle(self, checked):
        self.multi_box.setVisible(checked)
        if checked:
            # adopt the currently loaded image as the first slot (non-destructive)
            has_image = self.data.signal_full is not None or self.data.is_synthetic
            if has_image and self.data not in self.datasets:
                self.datasets.append(self.data)
                self.active_idx = len(self.datasets) - 1
            self._refresh_image_combo()
            self.status.showMessage("Multi-image mode on. Each loaded sample keeps its own "
                                    "reference; use the dropdown to switch.")
        else:
            self.status.showMessage("Multi-image mode off (current image stays active).")

    def on_image_combo_change(self, idx):
        if not (0 <= idx < len(self.datasets)):
            return
        self.active_idx = idx
        self.data = self.datasets[idx]
        self._restore_ui_for_active()
        # redisplay this image's already-processed phasor without recomputing
        self.phasor.set_data(self.data)
        self.last_overlay = None
        self.cluster_stats = []
        if self.chk_live.isChecked() and self.mode == "cursor" and self.phasor.cursors \
                and self.data.real_cal is not None:
            self._compute_cursor()                  # repaint shared cursors on this image
        else:
            self._fill_table()
            self.chk_overlay.blockSignals(True); self.chk_overlay.setChecked(False); self.chk_overlay.blockSignals(False)
            self.refresh_image()
        self.status.showMessage(f"Switched to image {idx + 1}: "
                                f"{os.path.basename(self.data.sample_path) or 'synthetic'}")

    def remove_image(self):
        if not (0 <= self.active_idx < len(self.datasets)):
            return
        self.datasets.pop(self.active_idx)
        if not self.datasets:
            self.active_idx = -1
            self._refresh_image_combo()
            return
        self.active_idx = min(self.active_idx, len(self.datasets) - 1)
        self._refresh_image_combo()
        self.on_image_combo_change(self.active_idx)

    def on_channel_change(self, idx):
        if self.data.signal_full is None and not self.data.is_synthetic:
            return
        self.data.channel = max(0, idx)
        self.apply_processing()

    def on_ref_channel_change(self, idx):
        if not self.data.ref_path:
            return
        self.data.ref_channel = max(0, min(idx, self.data.ref_n_channels - 1))
        if self.data.signal_full is not None or self.data.is_synthetic:
            self.apply_processing()

    # ---- processing --------------------------------------------------------
    def apply_processing(self):
        if self.data.signal_full is None and not self.data.is_synthetic:
            QtWidgets.QMessageBox.information(self, "No data", "Load a sample first.")
            return
        mode = self.cb_filter.currentText()
        if mode == "pawflim" and self.data.is_synthetic:
            QtWidgets.QMessageBox.information(
                self, "pawFLIM",
                "pawFLIM needs a real TCSPC signal (multi-harmonic) and isn't available "
                "for the synthetic demo. Using 'median' instead.")
            mode = "median"; self.cb_filter.setCurrentText("median")
        self.data.harmonic = self.sp_harm.value()
        self.data.frequency = self.sp_freq.value()
        self.data.channel = max(0, self.cb_channel.currentIndex())
        if self.data.ref_path:
            self.data.ref_channel = max(0, self.cb_ref_channel.currentIndex())
        t0 = time.perf_counter()
        try:
            self.data.apply_processing(
                ref_path=self.data.ref_path or None,
                ref_lifetime=self.sp_reflt.value(),
                filter_mode=mode,
                median_size=self.sp_msize.value(),
                median_repeat=self.sp_mrep.value(),
                paw_sigma=self.sp_psigma.value(),
                paw_levels=self.sp_plevels.value(),
                intensity_min=float(self.sp_thr.value()),
            )
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Processing error", repr(e)); return
        elapsed = time.perf_counter() - t0
        self.phasor.set_data(self.data)
        self.chk_overlay.blockSignals(True); self.chk_overlay.setChecked(False); self.chk_overlay.blockSignals(False)
        self.refresh_image()
        if self.data.ref_path:
            ref_note = f", ref ch {self.data.ref_channel}"
        else:
            ref_note = ""
        st = getattr(self.data, "_intensity_stats", {})
        if st:
            self.lbl_photon_range.setText(
                f"Image photon counts: {st['min']:.0f} – {st['max']:.0f} "
                f"(median {st['median']:.0f})")
            n_below = int(round(st["masked_pct"] * st.get("n_pixels", 0) / 100.0))
            int_msg = (f" | min photons ≥ {st['threshold']:.0f} "
                       f"({n_below} px removed)")
        else:
            int_msg = ""
        self.status.showMessage(
            f"Phasor recomputed — sample ch {self.data.channel}{ref_note}, "
            f"filter={mode}, H={self.data.harmonic}{int_msg}  ({self._fmt_elapsed(elapsed)})")

    # ---- cursor actions ----------------------------------------------------
    def add_cursor(self): self.phasor.add_cursor(radius=self.sld_radius.value() * 0.001)
    def remove_cursor(self):
        if self.mode != "cursor":
            return
        if self.phasor.selected < 0:
            self.status.showMessage("Click a circle on the phasor plot to select it, then Delete.")
            return
        self.phasor.remove_selected()
        self._refresh_after_cursor_edit()
        self.status.showMessage("Circle removed.")

    def clear_cursors(self):
        self.phasor.clear_cursors()
        self._refresh_after_cursor_edit()
        self.status.showMessage("All circles cleared.")

    def _refresh_after_cursor_edit(self):
        """Update segmentation overlay and table whenever circles are added/removed/moved."""
        self._paint_timer.stop()
        if self.mode != "cursor" or self.data.real_cal is None:
            return
        if self.phasor.cursors:
            self._compute_cursor()
        else:
            self.last_overlay = None
            self.cluster_stats = []
            self._fill_table()
            self.chk_overlay.blockSignals(True)
            self.chk_overlay.setChecked(False)
            self.chk_overlay.blockSignals(False)
            self.refresh_image()

    def _on_phasor_key_press(self, event):
        if self.mode != "cursor":
            return
        if event.key in ("delete", "backspace"):
            self.remove_cursor()

    def clear_gmm(self):
        self.phasor.clear_gmm()
        if hasattr(self, "gmm"):
            del self.gmm
        self.last_overlay = None
        self.cluster_stats = []
        self._fill_table()
        self.refresh_image()
        self.status.showMessage("GMM fit cleared.")

    def on_radius_slider(self, v):
        r = v * 0.001; self.lbl_radius.setText(f"{r:.3f}"); self.phasor.set_selected_radius(r)

    def _sync_radius_slider(self):
        i = self.phasor.selected
        if 0 <= i < len(self.phasor.cursors):
            r = self.phasor.cursors[i]["radius"]
            self.sld_radius.blockSignals(True); self.sld_radius.setValue(int(round(r * 1000))); self.sld_radius.blockSignals(False)
            self.lbl_radius.setText(f"{r:.3f}")

    def _live_active(self):
        return self.chk_live.isChecked() and self.mode == "cursor" and self.data.real_cal is not None

    def on_cursor_moving(self):
        """Interactive drag/scroll/slider: repaint overlay only (fast), defer the rest."""
        self._sync_radius_slider()
        if not self._live_active() or not self.phasor.cursors:
            return
        masks, colors = self._cursor_masks_colors()
        if masks is None:
            return
        intensity = self._segmentation_intensity()
        overlay = pseudo_color(*[masks[k] for k in range(len(masks))],
                               intensity=intensity, colors=np.array(colors))
        self.last_overlay = np.clip(np.asarray(overlay), 0, 1)
        if not self.chk_overlay.isChecked():
            self.chk_overlay.blockSignals(True); self.chk_overlay.setChecked(True); self.chk_overlay.blockSignals(False)
        self.image.update_overlay(self.last_overlay, title=f"Segmentation ({self.mode})")
        self._paint_timer.start()   # full recompute (lifetimes + table) once interaction settles

    def on_cursor_changed(self):
        """Committed change: sync slider and refresh segmentation (always, not only live-paint)."""
        self._sync_radius_slider()
        self._refresh_after_cursor_edit()

    def _deferred_full_compute(self):
        if self._live_active() and self.phasor.cursors:
            self._compute_cursor()

    # ---- GMM ---------------------------------------------------------------
    def fit_gmm(self):
        if not HAVE_SKLEARN:
            QtWidgets.QMessageBox.warning(self, "Missing dependency", "pip install scikit-learn"); return
        if not self.rb_gmm.isChecked():
            self.rb_gmm.setChecked(True)
        if self.data.real_cal is None:
            QtWidgets.QMessageBox.information(self, "GMM", "Load data and click Apply first."); return
        m = self.data.valid_mask()
        if m.sum() < 10:
            QtWidgets.QMessageBox.information(self, "GMM", "Not enough valid pixels."); return
        X = np.column_stack([self.data.real_cal[m], self.data.imag_cal[m]])
        cov = self.cb_cov.currentText()
        try:
            gm = self._fit_gmm_model(X, cov)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "GMM fit failed", str(e)); return
        self.gmm = gm
        covs = self._gmm_covariances(gm)
        colors = [categorical_rgb(k) for k in range(gm.n_components)]
        self.phasor.show_gmm(gm.means_, covs, colors)
        self._compute_gmm()
        return

    def _fit_gmm_model(self, X, cov):
        if self.chk_bic.isChecked():
            best, best_bic, best_n = None, np.inf, 1
            for n in range(1, self.sp_ncomp.value() + 1):
                gm = GaussianMixture(n, covariance_type=cov, random_state=0).fit(X)
                b = gm.bic(X)
                if b < best_bic:
                    best, best_bic, best_n = gm, b, n
            gm = best
            self.status.showMessage(f"GMM auto-selected {best_n} components (BIC={best_bic:.0f}).")
            return gm
        gm = GaussianMixture(self.sp_ncomp.value(), covariance_type=cov, random_state=0).fit(X)
        self.status.showMessage(f"GMM fit with {gm.n_components} components.")
        return gm

    def _gmm_covariances(self, gm):
        n, ct = gm.n_components, gm.covariance_type
        if ct == "full": return [gm.covariances_[k] for k in range(n)]
        if ct == "tied": return [gm.covariances_ for _ in range(n)]
        if ct == "diag": return [np.diag(gm.covariances_[k]) for k in range(n)]
        return [np.eye(2) * gm.covariances_[k] for k in range(n)]

    # ---- compute lifetimes + paint ----------------------------------------
    def compute_and_paint(self):
        if self.data.real_cal is None:
            return
        if self.mode == "cursor":
            self._compute_cursor()
        else:
            self._compute_gmm()

    def _lifetimes(self, g, s):
        freq = self.data.work_frequency
        tp, tm = phasor_to_apparent_lifetime(g, s, freq)
        tn = phasor_to_normal_lifetime(g, s, freq)
        return float(tp), float(tm), float(tn)

    def _cursor_masks_colors(self):
        """Boolean masks (within valid pixels) and colors for the current circles."""
        cur = self.phasor.cursors
        if not cur or self.data.real_cal is None:
            return None, None
        g, s = self.data.real_cal, self.data.imag_cal
        cr = np.array([c["center_real"] for c in cur])
        ci = np.array([c["center_imag"] for c in cur])
        rad = np.array([c["radius"] for c in cur])
        masks = mask_from_circular_cursor(g, s, cr, ci, radius=rad)
        if masks.ndim == 2:
            masks = masks[np.newaxis]
        masks = masks & self.data.valid_mask()[np.newaxis]
        return masks, [c["color"] for c in cur]

    def _compute_cursor(self):
        cur = self.phasor.cursors
        if not cur:
            QtWidgets.QMessageBox.information(self, "Cursors", "Add at least one circle."); return
        g, s = self.data.real_cal, self.data.imag_cal
        masks, colors = self._cursor_masks_colors()
        valid = self.data.valid_mask()
        total_valid = max(int(valid.sum()), 1)
        self.cluster_stats = []
        for k, c in enumerate(cur):
            mk = masks[k]; n = int(mk.sum())
            if n > 0:
                cg = float(np.nanmean(g[mk])); cs = float(np.nanmean(s[mk]))
                tp, tm, tn = self._lifetimes(cg, cs)
            else:
                cg = cs = tp = tm = tn = float("nan")
            self.cluster_stats.append(dict(idx=k + 1, color=c["color"], label=c["label"],
                                           tp=tp, tm=tm, tn=tn, g=cg, s=cs, n=n,
                                           area=100.0 * n / total_valid))
        self._paint(masks, colors); self._fill_table()

    def _compute_gmm(self):
        if not hasattr(self, "gmm"):
            QtWidgets.QMessageBox.information(self, "GMM", "Fit a GMM first."); return
        gm = self.gmm
        g, s = self.data.real_cal, self.data.imag_cal
        valid = self.data.valid_mask()
        if valid.sum() < 10:
            QtWidgets.QMessageBox.information(self, "GMM", "Not enough valid pixels."); return
        X = np.column_stack([g[valid], s[valid]])
        lab = gm.predict(X)
        n_comp = gm.n_components
        labelmap = np.full(g.shape, -1, dtype=int)
        labelmap[valid] = lab
        masks = np.stack([(labelmap == k) & valid for k in range(n_comp)])
        colors = [categorical_rgb(k) for k in range(n_comp)]
        total_valid = max(int(valid.sum()), 1)
        self.cluster_stats = []
        for k in range(n_comp):
            cg, cs = gm.means_[k]
            tp, tm, tn = self._lifetimes(float(cg), float(cs))
            n = int(masks[k].sum())
            self.cluster_stats.append(dict(idx=k + 1, color=colors[k], label=f"cluster {k + 1}",
                                           tp=tp, tm=tm, tn=tn, g=float(cg), s=float(cs), n=n,
                                           area=100.0 * n / total_valid))
        self._paint(masks, colors); self._fill_table()

    def _segmentation_intensity(self):
        """Background intensity for pseudo_color: raw photons, zero outside valid mask."""
        raw = self.data.mean_raw if self.data.mean_raw is not None else self.data.mean_thr
        if raw is None:
            return np.zeros((2, 2), dtype=np.float64)
        valid = self.data.valid_mask()
        out = np.zeros(np.shape(raw), dtype=np.float64)
        out[valid] = np.nan_to_num(raw[valid])
        return out

    def _paint(self, masks, colors):
        intensity = self._segmentation_intensity()
        overlay = pseudo_color(*[masks[k] for k in range(len(masks))],
                               intensity=intensity, colors=np.array(colors))
        self.last_overlay = np.clip(np.asarray(overlay), 0, 1)
        self.chk_overlay.setChecked(True); self.refresh_image()

    def refresh_image(self):
        if self.chk_overlay.isChecked() and self.last_overlay is not None:
            self.image.show_overlay(self.last_overlay, title=f"Segmentation ({self.mode})")
        else:
            self._show_base_image()

    def _show_base_image(self):
        """Render the currently selected base view (intensity or a lifetime map)."""
        choice = self.cb_base.currentText() if hasattr(self, "cb_base") else "Intensity"
        tau_map = {
            "Apparent \u03c4\u03c6 (ns)": self.data.tau_phi,
            "Apparent \u03c4m (ns)": self.data.tau_mod,
            "Normal \u03c4 (ns)": self.data.tau_normal,
        }.get(choice)
        if tau_map is not None:
            valid = self.data.valid_mask() if self.data.real_cal is not None else None
            disp = np.where(valid, tau_map, np.nan) if valid is not None else tau_map
            vmin, vmax = self._robust_range(disp)
            self.image.show_map(disp, title=choice, cmap="turbo", label="ns",
                                vmin=vmin, vmax=vmax)
        elif self.data.mean_raw is not None:
            self.image.show_intensity(self.data.mean_raw)
        elif self.data.mean_thr is not None:
            self.image.show_intensity(self.data.mean_thr)

    @staticmethod
    def _robust_range(arr):
        finite = np.asarray(arr)[np.isfinite(arr)]
        if finite.size == 0:
            return None, None
        return float(np.nanpercentile(finite, 2)), float(np.nanpercentile(finite, 98))

    # ---- results table -----------------------------------------------------
    def _fill_table(self):
        if getattr(self, "_label_signal_connected", False):
            self.table.itemChanged.disconnect(self._label_edited)
            self._label_signal_connected = False
        self.table.setRowCount(len(self.cluster_stats))
        for r, st in enumerate(self.cluster_stats):
            self.table.setItem(r, 0, self._ro(str(st["idx"])))
            sw = QtWidgets.QTableWidgetItem(); sw.setBackground(QtGui.QColor.fromRgbF(*st["color"]))
            sw.setFlags(Qt.ItemFlag.ItemIsEnabled); self.table.setItem(r, 1, sw)
            lab = QtWidgets.QTableWidgetItem(st["label"])
            lab.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsEditable | Qt.ItemFlag.ItemIsSelectable)
            self.table.setItem(r, 2, lab)
            self.table.setItem(r, 3, self._ro(f"{st['g']:.4f}"))
            self.table.setItem(r, 4, self._ro(f"{st['s']:.4f}"))
            self.table.setItem(r, 5, self._ro(f"{st['tp']:.3f}"))
            self.table.setItem(r, 6, self._ro(f"{st['tm']:.3f}"))
            self.table.setItem(r, 7, self._ro(f"{st['tn']:.3f}"))
            self.table.setItem(r, 8, self._ro(str(st["n"])))
            self.table.setItem(r, 9, self._ro(f"{st['area']:.2f}"))
        self.table.itemChanged.connect(self._label_edited)
        self._label_signal_connected = True

    def _label_edited(self, item):
        if item.column() == 2:
            r = item.row()
            if 0 <= r < len(self.cluster_stats):
                self.cluster_stats[r]["label"] = item.text()
                if self.mode == "cursor" and r < len(self.phasor.cursors):
                    self.phasor.cursors[r]["label"] = item.text()

    @staticmethod
    def _ro(text):
        it = QtWidgets.QTableWidgetItem(text); it.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable); return it

    @staticmethod
    def _rgb_hex(c):
        return "FF" + "".join(f"{int(round(255 * max(0.0, min(1.0, x)))):02X}" for x in c[:3])

    # ---- export ------------------------------------------------------------
    def save_overlay(self):
        if self.last_overlay is None: return
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Save overlay", "segmentation.png", "PNG (*.png)")
        if path:
            import matplotlib.pyplot as plt
            plt.imsave(path, self.last_overlay); self.status.showMessage(f"Saved {path}")

    def save_phasor(self):
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Save phasor", "phasor.png", "PNG (*.png)")
        if path:
            self.phasor.fig.savefig(path, dpi=200); self.status.showMessage(f"Saved {path}")

    def export_csv(self):
        if not self.cluster_stats: return
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Export results", "clusters.csv", "CSV (*.csv)")
        if not path: return
        with open(path, "w", newline="") as f:
            w = csv.writer(f)
            w.writerow(["cluster", "label", "g", "s", "tau_phi_ns", "tau_mod_ns", "tau_normal_ns",
                        "pixels", "area_percent", "frequency_MHz", "harmonic",
                        "sample_channel", "ref_channel", "filter", "sample", "reference"])
            for st in self.cluster_stats:
                w.writerow([st["idx"], st["label"], f"{st['g']:.5f}", f"{st['s']:.5f}",
                            f"{st['tp']:.4f}", f"{st['tm']:.4f}", f"{st['tn']:.4f}",
                            st["n"], f"{st['area']:.3f}",
                            f"{self.data.work_frequency:.4f}", self.data.harmonic,
                            self.data.channel,
                            self.data.ref_channel if self.data.ref_path else "",
                            self.cb_filter.currentText(),
                            os.path.basename(self.data.sample_path),
                            os.path.basename(self.data.ref_path) if self.data.ref_path else ""])
        self.status.showMessage(f"Exported {path}")

    def export_xlsx(self):
        if not self.cluster_stats:
            QtWidgets.QMessageBox.information(self, "Excel", "Nothing to export yet — compute clusters first.")
            return
        path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, "Export results (Excel)", "clusters.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            QtWidgets.QMessageBox.warning(
                self, "Missing dependency",
                "Excel export needs openpyxl:\n\n    pip install openpyxl")
            return
        try:
            import datetime
            wb = openpyxl.Workbook()
            ws = wb.active; ws.title = "Clusters"
            headers = ["#", "Color", "Label (what you see)", "g", "s",
                       "tau_phi (ns)", "tau_mod (ns)", "tau_normal (ns)",
                       "Pixels", "Area %"]
            ws.append(headers)
            for c in ws[1]:
                c.font = Font(bold=True); c.alignment = Alignment(horizontal="center")
            for st in self.cluster_stats:
                ws.append([st["idx"], "", st["label"],
                           round(st["g"], 5), round(st["s"], 5),
                           round(st["tp"], 4), round(st["tm"], 4), round(st["tn"], 4),
                           int(st["n"]), round(st["area"], 3)])
                ws.cell(row=ws.max_row, column=2).fill = PatternFill(
                    "solid", fgColor=self._rgb_hex(st["color"]))
            for i, w in enumerate([5, 8, 28, 10, 10, 12, 12, 14, 11, 9], start=1):
                ws.column_dimensions[get_column_letter(i)].width = w
            ws.freeze_panes = "A2"

            meta = wb.create_sheet("Metadata")
            meta.append(["Parameter", "Value"])
            for c in meta[1]:
                c.font = Font(bold=True)
            for k, v in [
                ("Sample", os.path.basename(self.data.sample_path)),
                ("Reference", os.path.basename(self.data.ref_path) if self.data.ref_path else ""),
                ("Laser frequency (MHz)", round(self.data.frequency, 4)),
                ("Working frequency (MHz)", round(self.data.work_frequency, 4)),
                ("Harmonic", self.data.harmonic),
                ("Sample channel", self.data.channel),
                ("Reference channel", self.data.ref_channel if self.data.ref_path else ""),
                ("Filter", self.cb_filter.currentText()),
                ("Segmentation mode", self.mode),
                ("Number of clusters", len(self.cluster_stats)),
                ("Exported", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ]:
                meta.append([k, v])
            meta.column_dimensions["A"].width = 26
            meta.column_dimensions["B"].width = 42
            wb.save(path)
            self.status.showMessage(f"Exported {path}")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Excel export error", repr(e))


def main():
    app = QtWidgets.QApplication(sys.argv)
    win = MainWindow(); win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
