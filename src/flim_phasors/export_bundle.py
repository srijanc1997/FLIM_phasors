"""Export a complete FLIM phasor analysis folder for sharing and archiving.

Writes phasor plots, per-sample lifetime and photon maps (PNG + ``maps.npz``),
cluster masks, CSV/Excel tables (g, s, τ_φ, τ_m, τ_n, cluster areas), and
``session.json`` with calibration, cursors, GMM fit, and cluster stats.
Expects a GUI ``MainWindow`` instance for live plot state and processing labels.
"""

from __future__ import annotations

import csv
import json
import os
import re
from datetime import datetime, timezone
from pathlib import Path

import numpy as np

from flim_phasors import __version__
from flim_phasors.analysis import cursor_masks_for_dataset, label_pixels_by_gmm, lifetimes_at_phasor
from flim_phasors.cursors_io import cursors_to_list
from flim_phasors.gui.processing import filter_label_for_dataset
from flim_phasors.session_bundle_io import (
    MAP_KEYS,
    _maps_from_dataset,
    _serialize_cluster_stats,
    _serialize_gmm_fit,
)
from flim_phasors.utils import (
    categorical_name,
    categorical_rgb,
    effective_reference_path,
    sample_core_metadata,
)


def _safe_name(text: str, max_len: int = 80) -> str:
    """Sanitize a string for use as a filesystem folder or file name.

    Replaces characters that are illegal in Windows paths (``<>:"/\\|?*``)
    with underscores, collapses runs of whitespace to a single underscore,
    and truncates to ``max_len`` so deeply nested export trees stay within
    path-length limits. An empty or all-sanitized input falls back to
    ``"sample"`` rather than producing an empty path segment.

    Args:
        text: Raw label or path fragment.
        max_len: Maximum returned length.

    Returns:
        Underscore-separated name safe for Windows paths.
    """
    text = (text or "sample").strip()
    text = re.sub(r'[<>:"/\\|?*]', "_", text)
    text = re.sub(r"\s+", "_", text)
    return text[:max_len] or "sample"


def _sample_stem(d, index: int = 0) -> str:
    """Return a filesystem-safe stem for the analyzed sample file.

    Uses ``display_name`` or the sample path basename (without extension).

    Args:
        d: ``PhasorData`` with path / display metadata.
        index: Zero-based index used when no path is available.

    Returns:
        Sanitized stem suitable as a filename prefix.
    """
    display = (getattr(d, "display_name", "") or "").strip()
    if display:
        return _safe_name(Path(display).stem)
    path = getattr(d, "sample_path", None) or ""
    if path:
        return _safe_name(Path(path).stem)
    return _safe_name(f"sample_{index + 1:02d}")


def _sample_folder_name(d, index: int) -> str:
    """Build a unique subfolder name for one exported sample.

    Every folder is prefixed with the sample's 1-based export index so
    folder names are always unique, even when two samples share a display
    name (e.g. two channels of the same file with no custom name). When a
    sample belongs to a group, the group name is inserted before the index
    (``{group}__{index}_{stem}``) so samples sharing a group still sort and
    scan together in a file browser.

    Args:
        d: ``PhasorData`` with optional ``group_name`` and path metadata.
        index: Zero-based sample index in the session.

    Returns:
        Sanitized folder name, always unique per export index.
    """
    base = _sample_stem(d, index)
    group = (getattr(d, "group_name", "") or "").strip()
    indexed = f"{index + 1:02d}_{base}"
    if group:
        return _safe_name(f"{group}__{indexed}")
    return _safe_name(indexed)


def _clear_dir(path: Path) -> None:
    """Remove all files in ``path`` (keeps subdirectories).

    Used so re-exporting the same session overwrites cleanly instead of leaving
    leftover images from older naming schemes.
    """
    if not path.is_dir():
        return
    for child in path.iterdir():
        if child.is_file():
            try:
                child.unlink()
            except OSError:
                pass


def _session_excel_path(out: Path) -> Path:
    """Return the fixed Excel workbook path for a session export folder.

    Using one stable filename (rather than one named after the active sample)
    means re-running Export all overwrites the previous workbook instead of
    accumulating a new file per export, and any legacy ``analysis_results.xlsx``
    from an older naming scheme can be detected and removed.

    Args:
        out: Root export directory for the session.

    Returns:
        Path ``out / "analysis.xlsx"``.
    """
    return out / "analysis.xlsx"


def _write_map_png(path: Path, arr, *, cmap="viridis", vmin=None, vmax=None):
    """Write a false-color map PNG with robust percentile scaling.

    NaN pixels (masked/invalid) are excluded from both the percentile
    computation and the rendered image via a masked array, so a small number
    of outlier bright/dark pixels does not wash out the visible dynamic
    range of the rest of the map. The PNG is written at native array
    resolution with no axes or colorbar; use
    :func:`_write_map_fig_with_colorbar` when a labeled figure is needed
    instead.

    Args:
        path: Output ``.png`` path.
        arr: 2-D numeric array (NaN masked).
        cmap: Matplotlib colormap name.
        vmin: Optional lower display bound; 2nd percentile if None.
        vmax: Optional upper display bound; 98th percentile if None.

    Returns:
        False when no finite values exist; True after successful write.
    """
    import matplotlib.pyplot as plt

    data = np.asarray(arr, dtype=float)
    finite = data[np.isfinite(data)]
    if finite.size == 0:
        return False
    if vmin is None or vmax is None:
        lo, hi = np.percentile(finite, [2, 98])
        vmin = lo if vmin is None else vmin
        vmax = hi if vmax is None else vmax
    if vmax <= vmin:
        vmax = vmin + 1.0
    masked = np.ma.masked_invalid(data)
    plt.imsave(path, masked, cmap=cmap, vmin=vmin, vmax=vmax)
    return True


def _write_photon_png(path: Path, arr):
    """Write a grayscale photon-count map PNG with percentile contrast.

    Contrast bounds are fixed at the 2nd/98th percentile of finite values
    (not the raw min/max) so a handful of saturated or zero-count pixels do
    not compress the visible range for the rest of the field of view. Used
    for both the thresholded photon map and the brightfield (pre-threshold)
    map, which is why the colormap is always grayscale rather than
    parameterized like :func:`_write_map_png`.

    Args:
        path: Output ``.png`` path.
        arr: 2-D photon or intensity array.

    Returns:
        False when no finite values exist; True after successful write.
    """
    import matplotlib.pyplot as plt

    data = np.asarray(arr, dtype=float)
    finite = data[np.isfinite(data)]
    if finite.size == 0:
        return False
    lo, hi = np.percentile(finite, [2, 98])
    if hi <= lo:
        hi = lo + 1.0
    masked = np.ma.masked_invalid(data)
    plt.imsave(path, masked, cmap="gray", vmin=lo, vmax=hi)
    return True


def _write_map_fig_with_colorbar(
    path: Path,
    arr,
    *,
    title: str,
    cmap: str = "viridis",
    label: str = "ns",
) -> bool:
    """Write a τ (or similar) map figure PNG including a colorbar.

    Unlike :func:`_write_map_png`, this renders a full Matplotlib figure
    (with a title and a labeled colorbar) so the PNG is self-describing when
    viewed outside the app, e.g. in a shared report. Display bounds are
    fixed at the 2nd/98th percentile of finite values, matching the raw
    (colorbar-less) map so the two paired outputs stay visually consistent.

    Args:
        path: Output ``.png`` path (e.g. ``tau_mod_ns_fig.png``).
        arr: 2-D numeric array.
        title: Axes title.
        cmap: Matplotlib colormap name.
        label: Colorbar label.

    Returns:
        False when no finite values exist; True after successful write.
    """
    import matplotlib.pyplot as plt

    data = np.asarray(arr, dtype=float)
    finite = data[np.isfinite(data)]
    if finite.size == 0:
        return False
    lo, hi = np.percentile(finite, [2, 98])
    if hi <= lo:
        hi = lo + 1.0
    masked = np.ma.masked_invalid(data)
    fig, ax = plt.subplots(figsize=(5, 4), dpi=150)
    im = ax.imshow(masked, cmap=cmap, vmin=lo, vmax=hi)
    ax.set_title(title, fontsize=10)
    ax.axis("off")
    cbar = fig.colorbar(im, ax=ax, fraction=0.046, pad=0.04)
    cbar.set_label(label)
    fig.savefig(path, bbox_inches="tight")
    plt.close(fig)
    return True


def _write_maps_npz(sample_dir: Path, d, *, prefix: str) -> str | None:
    """Write quantitative phasor/lifetime arrays as ``{prefix}_maps.npz``.

    Delegates to :func:`~flim_phasors.session_bundle_io._maps_from_dataset`
    for the same map set used by session bundles (calibrated g/s, photon
    counts, and the three lifetime maps), so exported ``.npz`` files can be
    reloaded with plain NumPy for downstream analysis without needing this
    package installed. Nothing is written when the dataset has no computed
    maps yet.

    Args:
        sample_dir: Per-sample export directory.
        d: Processed ``PhasorData``.
        prefix: Sample filename stem used in the output basename.

    Returns:
        Basename when at least one array is written, else ``None``.
    """
    maps = _maps_from_dataset(d)
    if not maps:
        return None
    name = f"{prefix}_maps.npz"
    np.savez_compressed(sample_dir / name, **maps)
    return name


def cluster_rows_for_dataset(win, d, stats):
    """Build CSV row dicts for GMM or cursor cluster statistics on one sample.

    Each row records phasor center (g, s), apparent lifetimes, pixel count,
    area fraction, harmonic, reference path, and active filter settings.

    Args:
        win: Main window for reference path and filter resolution.
        d: ``PhasorData`` for the clustered sample.
        stats: Iterable of cluster stat dicts (``idx``, ``label``, ``g``, ``s``,
            ``tp``, ``tm``, ``tn``, ``n``, ``area``).

    Returns:
        List of row dictionaries keyed for ``clusters.csv`` columns.
    """
    rows = []
    # Shared ref checkbox overrides per-sample ref_path when resolving export metadata.
    ref = effective_reference_path(win, d)
    ref_base = os.path.basename(ref) if ref else ""
    sample_base = os.path.basename(d.sample_path) if d.sample_path else ""
    group = (getattr(d, "group_name", "") or "").strip()
    for st in stats:
        rows.append({
            "sample": sample_base,
            "group": group,
            "cluster": st["idx"],
            "label": st["label"],
            "g": st["g"],
            "s": st["s"],
            "tau_phi_ns": st["tp"],
            "tau_mod_ns": st["tm"],
            "tau_normal_ns": st["tn"],
            "pixels": st["n"],
            "area_percent": st["area"],
            "frequency_MHz": d.work_frequency,
            "harmonic": d.harmonic,
            "sample_channel": d.channel,
            "ref_channel": d.ref_channel if d.ref_path else "",
            "filter": filter_label_for_dataset(win, d),
            "reference": ref_base,
        })
    return rows


def sample_metadata_row(win, d, index: int) -> dict:
    """Build one row of sample-level metadata for summary CSV and session JSON."""
    ref = effective_reference_path(win, d)
    st = getattr(d, "_intensity_stats", {}) or {}
    core = sample_core_metadata(d, index, reference_path=ref)
    return {
        "index": index + 1,
        "label": core["label"],
        "sample_path": d.sample_path,
        "display_name": core["display_name"],
        "group": core["group"],
        "channel": core["channel"],
        "frequency_MHz": core["frequency_MHz"],
        "harmonic": core["harmonic"],
        "work_frequency_MHz": core["work_frequency_MHz"],
        "reference_path": core["reference_path"],
        "reference_channel": d.ref_channel if ref else "",
        "filter": filter_label_for_dataset(win, d),
        "min_photons": st.get("threshold", 0),
        "pixels_total": st.get("n_pixels", ""),
        "pixels_masked": st.get("n_below", ""),
        "computed": core["computed"],
    }


def build_session_dict(win) -> dict:
    """Serialize GUI session state for ``session.json`` in the export bundle."""
    datasets = win._all_datasets() if hasattr(win, "_all_datasets") else [win.data]
    cursors = cursors_to_list(list(win.phasor.cursors)) if hasattr(win, "phasor") else []
    try:
        import phasorpy
        pp_ver = getattr(phasorpy, "__version__", "")
    except ImportError:
        pp_ver = ""
    return {
        "app_version": __version__,
        "phasorpy_version": pp_ver,
        "exported_utc": datetime.now(timezone.utc).isoformat(),
        "segmentation_mode": getattr(win, "mode", ""),
        "shared_reference": bool(win.chk_shared_ref.isChecked()) if hasattr(win, "chk_shared_ref") else False,
        "shared_reference_path": getattr(win, "shared_ref_path", ""),
        "calibration": {
            # When shared_reference is true, GUI uses shared_reference_path for all samples.
            "frequency_MHz": win.sp_freq.value(),
            "harmonic": win.sp_harm.value(),
            "reference_lifetime_ns": win.sp_reflt.value(),
            "filter": filter_label_for_dataset(win, win.data),
            "min_photons": win.sp_thr.value(),
            "harmonic_mask": win.chk_detect_harm.isChecked() if hasattr(win, "chk_detect_harm") else True,
            "reference_path": getattr(win, "shared_ref_path", "") or getattr(win.data, "ref_path", ""),
            "reference_channel": getattr(win, "shared_ref_channel", 0),
            "mean_g": getattr(win.ref_calibration, "mean_g", 0.0) if hasattr(win, "ref_calibration") else 0.0,
            "mean_s": getattr(win.ref_calibration, "mean_s", 0.0) if hasattr(win, "ref_calibration") else 0.0,
            "harmonic_gs": (
                [[float(g), float(s)] for g, s in (win.ref_calibration.harmonic_gs or [])]
                if hasattr(win, "ref_calibration") and getattr(win.ref_calibration, "harmonic_gs", None)
                else None
            ),
            "manual": bool(getattr(win.ref_calibration, "use_manual", False)) if hasattr(win, "ref_calibration") else False,
            "manual_g": getattr(win.ref_calibration, "manual_g", 0.0) if hasattr(win, "ref_calibration") else 0.0,
            "manual_s": getattr(win.ref_calibration, "manual_s", 0.0) if hasattr(win, "ref_calibration") else 0.0,
        },
        "samples": [sample_metadata_row(win, d, i) for i, d in enumerate(datasets)],
        "active_sample_index": getattr(win, "active_idx", -1),
        "cursors": cursors,
        "gmm_fit": _serialize_gmm_fit(getattr(win, "_gmm_fit", None)),
        "cluster_stats": _serialize_cluster_stats(getattr(win, "cluster_stats", [])),
        "maps_npz_keys": list(MAP_KEYS),
    }


def write_dict_csv(path: Path, rows: list[dict]):
    """Write homogeneous row dicts to a UTF-8 CSV (no-op when ``rows`` is empty)."""
    if not rows:
        return
    fields = list(rows[0].keys())
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)


def write_clusters_csv(path: Path, rows: list[dict]):
    """Write cluster statistics rows to a UTF-8 CSV file."""
    write_dict_csv(path, rows)


def write_samples_summary(path: Path, rows: list[dict]):
    """Write per-sample metadata rows to a UTF-8 CSV summary file."""
    write_dict_csv(path, rows)


def write_excel_bundle(path: Path, win, all_cluster_rows: list[dict], sample_rows: list[dict]):
    """Write a multi-sheet Excel workbook with summary, samples, and clusters.

    Cluster label cells are filled with categorical colors by cluster index.
    Requires ``openpyxl``; callers should catch ``ImportError``.

    Args:
        path: Destination ``.xlsx`` path.
        win: Main window for export metadata.
        all_cluster_rows: Combined cluster CSV rows (may be empty).
        sample_rows: Per-sample metadata rows for the Samples sheet.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.append(["Parameter", "Value"])
    for c in summary[1]:
        c.font = Font(bold=True)
    meta = [
        ("Software", f"FLIM Phasors {__version__}"),
        ("Exported (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")),
        ("Samples", len(sample_rows)),
        ("Segmentation mode", getattr(win, "mode", "")),
        ("Filter", filter_label_for_dataset(win, win.data)),
    ]
    for k, v in meta:
        summary.append([k, v])
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 48

    ws_samples = wb.create_sheet("Samples")
    if sample_rows:
        headers = list(sample_rows[0].keys())
        ws_samples.append(headers)
        for c in ws_samples[1]:
            c.font = Font(bold=True)
        for row in sample_rows:
            ws_samples.append([row[h] for h in headers])

    if all_cluster_rows:
        ws_cl = wb.create_sheet("Clusters")
        headers = list(all_cluster_rows[0].keys())
        ws_cl.append(headers)
        for c in ws_cl[1]:
            c.font = Font(bold=True)
        for row in all_cluster_rows:
            ws_cl.append([row[h] for h in headers])

        if hasattr(win, "_rgb_hex"):
            col_idx = headers.index("label") + 1
            for i, row in enumerate(all_cluster_rows, start=2):
                try:
                    k = int(row.get("cluster", 1)) - 1
                except (TypeError, ValueError):
                    k = 0
                color = categorical_rgb(max(0, k))
                ws_cl.cell(row=i, column=col_idx).fill = PatternFill(
                    "solid", fgColor=win._rgb_hex(color))

    wb.save(path)


def cluster_stats_and_masks_for_dataset(win, d, *, is_active: bool = False):
    """Compute cluster statistics and boolean masks for one sample.

    Prefers results stored on the dataset (``d.gmm_fit`` / ``d.cluster_stats``)
    so Export all can save every sample's remembered GMM after switching images.
    Falls back to the active window fit/cursors when a sample has no stash.

    Args:
        win: Main window (mode, cursors, GMM fit, optional active stats).
        d: Computed ``PhasorData``.
        is_active: True when ``d`` is the GUI active sample.

    Returns:
        Tuple ``(stats, masks)`` where ``stats`` is a list of cluster dicts and
        ``masks`` is ``(n, H, W)`` bool array or ``None``.
    """
    if d.real_cal is None:
        return [], None

    g, s = d.real_cal, d.imag_cal
    valid = d.valid_mask()
    total_valid = max(int(valid.sum()), 1)
    freq = float(d.work_frequency)

    stored_fit = getattr(d, "gmm_fit", None)
    stored_stats = list(getattr(d, "cluster_stats", None) or [])

    # Per-sample stash wins so Export all captures GMM after switching images.
    if stored_fit is not None:
        cr, ci, rm, _ri, _ang = stored_fit
        labelmap = label_pixels_by_gmm(g, s, cr, ci, rm)
        n_comp = len(cr)
        masks = np.stack([(labelmap == k) & valid for k in range(n_comp)])
        if stored_stats:
            return stored_stats, masks
        stats = []
        for k in range(n_comp):
            cg, cs = float(cr[k]), float(ci[k])
            mk = masks[k]
            tp, tm, tn = lifetimes_at_phasor(cg, cs, freq)
            n = int(mk.sum())
            stats.append(dict(
                idx=k + 1, color=categorical_rgb(k), label=categorical_name(k),
                tp=tp, tm=tm, tn=tn, g=cg, s=cs, n=n,
                area=100.0 * n / total_valid,
            ))
        return stats, masks

    mode = getattr(win, "mode", "cursor")
    if mode == "gmm" and hasattr(win, "_gmm_fit") and win._gmm_fit is not None:
        cr, ci, rm, _ri, _ang = win._gmm_fit
        labelmap = label_pixels_by_gmm(g, s, cr, ci, rm)
        n_comp = len(cr)
        masks = np.stack([(labelmap == k) & valid for k in range(n_comp)])
        if is_active and getattr(win, "cluster_stats", None):
            return list(win.cluster_stats), masks
        stats = []
        for k in range(n_comp):
            cg, cs = float(cr[k]), float(ci[k])
            mk = masks[k]
            tp, tm, tn = lifetimes_at_phasor(cg, cs, freq)
            n = int(mk.sum())
            stats.append(dict(
                idx=k + 1, color=categorical_rgb(k), label=categorical_name(k),
                tp=tp, tm=tm, tn=tn, g=cg, s=cs, n=n,
                area=100.0 * n / total_valid,
            ))
        return stats, masks

    if stored_stats and mode == "cursor":
        cursors = getattr(getattr(win, "phasor", None), "cursors", None) or []
        masks = cursor_masks_for_dataset(d, cursors)
        return stored_stats, masks

    cursors = getattr(getattr(win, "phasor", None), "cursors", None) or []
    masks = cursor_masks_for_dataset(d, cursors)
    if masks is None:
        if is_active and getattr(win, "cluster_stats", None):
            return list(win.cluster_stats), None
        return [], None

    if is_active and getattr(win, "cluster_stats", None):
        return list(win.cluster_stats), masks

    stats = []
    for k, c in enumerate(cursors):
        mk = masks[k]
        n = int(mk.sum())
        if n > 0:
            cg = float(np.nanmean(g[mk]))
            cs = float(np.nanmean(s[mk]))
            tp, tm, tn = lifetimes_at_phasor(cg, cs, freq)
        else:
            cg = cs = tp = tm = tn = float("nan")
        stats.append(dict(
            idx=k + 1,
            color=c.get("color", categorical_rgb(k)),
            label=c.get("label", categorical_name(k)),
            tp=tp, tm=tm, tn=tn, g=cg, s=cs, n=n,
            area=100.0 * n / total_valid,
        ))
    return stats, masks


def _export_cluster_masks(
    sample_dir: Path, masks: np.ndarray, written: list[str], *, prefix: str,
):
    """Write ``{prefix}_mask_01.png`` … grayscale masks for each cluster.

    Each boolean slice along the first axis of ``masks`` is scaled to
    ``0``/``255`` and written as an independent single-channel PNG, numbered
    from 1 in cluster order (matching the ``idx`` field used in cluster stat
    rows), so masks can be loaded individually in external tools like
    ImageJ/Fiji without needing to know the original cluster count.

    Args:
        sample_dir: Per-sample export subdirectory.
        masks: Boolean array ``(n, H, W)``.
        written: Mutable list of output paths appended in place.
        prefix: Sample filename stem.
    """
    import matplotlib.pyplot as plt

    for k in range(masks.shape[0]):
        name = f"{prefix}_mask_{k + 1:02d}.png"
        path = sample_dir / name
        plt.imsave(path, masks[k].astype(np.uint8) * 255, cmap="gray", vmin=0, vmax=255)
        written.append(str(path))


def export_sample_maps(sample_dir: Path, d, *, prefix: str | None = None, index: int = 0):
    """Write lifetime/photon PNGs, colorbar pairs, and ``maps.npz`` for one sample.

    Each heatmap and brightfield is saved twice: without colorbar (raster) and
    with colorbar (``*_colorbar.png``). Filenames are prefixed with the analyzed
    sample stem.

    Args:
        sample_dir: Directory created under ``samples/`` in the bundle.
        d: Processed ``PhasorData`` with lifetime maps.
        prefix: Optional filename stem; derived from ``d`` when omitted.
        index: Sample index used when deriving the stem.

    Returns:
        List of basenames of maps successfully written.
    """
    prefix = prefix or _sample_stem(d, index)
    written = []

    npz_name = _write_maps_npz(sample_dir, d, prefix=prefix)
    if npz_name:
        written.append(npz_name)

    # Masked / thresholded photons (mean_thr: NaN where g/s invalid or below Min N)
    thr = d.mean_thr if d.mean_thr is not None else d.mean_raw
    if thr is not None:
        name = f"{prefix}_photons.png"
        if _write_photon_png(sample_dir / name, thr):
            written.append(name)
        cbar_name = f"{prefix}_photons_colorbar.png"
        if _write_map_fig_with_colorbar(
            sample_dir / cbar_name, thr, title=f"{prefix} photons (masked)",
            cmap="gray", label="counts",
        ):
            written.append(cbar_name)

    # Brightfield: pre-threshold counts (LIF photon image or raw histogram sum).
    bright = None
    if hasattr(d, "intensity_brightfield"):
        bright = d.intensity_brightfield()
    if bright is None:
        bright = d.mean_raw
    if bright is not None:
        name = f"{prefix}_brightfield.png"
        if _write_photon_png(sample_dir / name, bright):
            written.append(name)
        cbar_name = f"{prefix}_brightfield_colorbar.png"
        if _write_map_fig_with_colorbar(
            sample_dir / cbar_name, bright, title=f"{prefix} brightfield",
            cmap="gray", label="counts",
        ):
            written.append(cbar_name)

    tau_maps = [
        ("tau_phi_ns", d.tau_phi, "τφ (ns)"),
        ("tau_mod_ns", d.tau_mod, "τmod (ns)"),
        ("tau_normal_ns", d.tau_normal, "τ normal (ns)"),
    ]
    for stem, arr, title in tau_maps:
        if arr is None:
            continue
        raster = f"{prefix}_{stem}.png"
        if _write_map_png(sample_dir / raster, arr, cmap="viridis"):
            written.append(raster)
        cbar_name = f"{prefix}_{stem}_colorbar.png"
        if _write_map_fig_with_colorbar(
            sample_dir / cbar_name, arr, title=f"{prefix} {title}",
            cmap="viridis", label=title,
        ):
            written.append(cbar_name)

    return written


def export_analysis_bundle(win, out_dir: str | Path) -> dict:
    """Write a complete FLIM phasor analysis export folder.

    Creates phasor figure exports (PNG/PDF/SVG), active segmentation overlay,
    per-sample map subfolders (PNG + ``maps.npz`` + masks), CSV tables for all
    computed samples, optional Excel workbook, session JSON (including GMM and
    cluster stats), and a README describing contents.

    Args:
        win: ``MainWindow`` with phasor canvas, cluster stats, and datasets.
        out_dir: Destination directory (created if missing).

    Returns:
        Log dict with ``directory``, ``files`` (paths written), and ``n_samples``.

    Raises:
        ValueError: When no loaded sample is available to export.
    """
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)

    written: list[str] = []
    datasets = win._all_datasets() if hasattr(win, "_all_datasets") else [win.data]
    datasets = [
        d for d in datasets
        if d.sample_path or d.signal_full is not None or d.real_cal is not None
    ]
    if not datasets:
        raise ValueError("No loaded sample to export.")
    for d in datasets:
        if d.signal_full is None and d.sample_path and hasattr(d, "ensure_loaded"):
            try:
                # session.json restore has paths only; try decode before skipping export.
                d.ensure_loaded(frame=getattr(d, "frame_index", -1))
            except Exception:
                pass

    active = getattr(win, "data", None)
    active_stem = _sample_stem(active, getattr(win, "active_idx", 0) or 0) if active else "session"

    # Phasor plot named for the active sample (overwrite on re-export)
    if hasattr(win, "phasor") and getattr(win.phasor, "fig", None) is not None:
        for obsolete in ("phasor_plot.png", "phasor_plot.pdf", "phasor_plot.svg"):
            old = out / obsolete
            if old.is_file():
                try:
                    old.unlink()
                except OSError:
                    pass
        for ext, fmt in (
            (f"{active_stem}_phasor.png", "png"),
            (f"{active_stem}_phasor.pdf", "pdf"),
            (f"{active_stem}_phasor.svg", "svg"),
        ):
            try:
                p = out / ext
                kwargs = {"bbox_inches": "tight"}
                if fmt == "png":
                    kwargs["dpi"] = 200
                else:
                    kwargs["format"] = fmt
                win.phasor.fig.savefig(p, **kwargs)
                written.append(str(p))
            except Exception:
                pass

    # Active segmentation overlay (root) + per-sample overlays
    if hasattr(win, "_stash_segmentation_to_dataset"):
        try:
            win._stash_segmentation_to_dataset(win.data)
        except Exception:
            pass

    if getattr(win, "last_overlay", None) is not None:
        import matplotlib.pyplot as plt

        old = out / "segmentation_active.png"
        if old.is_file():
            try:
                old.unlink()
            except OSError:
                pass
        seg_path = out / f"{active_stem}_segmentation.png"
        plt.imsave(seg_path, np.clip(np.asarray(win.last_overlay), 0, 1))
        written.append(str(seg_path))

    # Per-sample subfolders (cleared so re-exports do not accumulate duplicates)
    sample_rows = []
    all_cluster_rows: list[dict] = []
    samples_root = out / "samples"
    samples_root.mkdir(parents=True, exist_ok=True)

    for i, d in enumerate(datasets):
        sample_rows.append(sample_metadata_row(win, d, i))
        if d.real_cal is None:
            continue
        prefix = _sample_stem(d, i)
        sub = samples_root / _sample_folder_name(d, i)
        sub.mkdir(parents=True, exist_ok=True)
        _clear_dir(sub)
        maps = export_sample_maps(sub, d, prefix=prefix, index=i)
        written.extend(str(sub / m) for m in maps)

        is_active = d is active
        stats, masks = cluster_stats_and_masks_for_dataset(win, d, is_active=is_active)
        if stats:
            rows = cluster_rows_for_dataset(win, d, stats)
            all_cluster_rows.extend(rows)
            csv_name = f"{prefix}_clusters.csv"
            write_clusters_csv(sub / csv_name, rows)
            written.append(str(sub / csv_name))
        if masks is not None and masks.size:
            _export_cluster_masks(sub, masks, written, prefix=prefix)
        overlay = getattr(d, "last_overlay", None)
        if overlay is not None:
            import matplotlib.pyplot as plt
            seg_name = f"{prefix}_segmentation.png"
            plt.imsave(sub / seg_name, np.clip(np.asarray(overlay), 0, 1))
            written.append(str(sub / seg_name))

    # Tables (always overwrite)
    write_samples_summary(out / "samples_summary.csv", sample_rows)
    written.append(str(out / "samples_summary.csv"))

    if all_cluster_rows:
        write_clusters_csv(out / "clusters.csv", all_cluster_rows)
        written.append(str(out / "clusters.csv"))

    # One Excel workbook for the session — rewritten on every export
    xlsx_path = _session_excel_path(out)
    old_xlsx = out / "analysis_results.xlsx"
    if old_xlsx.is_file() and old_xlsx != xlsx_path:
        try:
            old_xlsx.unlink()
        except OSError:
            pass
    openpyxl_note = ""
    try:
        write_excel_bundle(xlsx_path, win, all_cluster_rows, sample_rows)
        written.append(str(xlsx_path))
    except ImportError:
        openpyxl_note = (
            "\nNote: Install openpyxl for Excel export: pip install openpyxl\n"
        )

    # Session JSON
    session_path = out / "session.json"
    session_path.write_text(
        json.dumps(build_session_dict(win), indent=2),
        encoding="utf-8",
    )
    written.append(str(session_path))

    keys_line = ", ".join(MAP_KEYS)
    readme = out / "README_export.txt"
    readme.write_text(
        f"FLIM Phasors export ({__version__})\n"
        f"Generated: {datetime.now().isoformat(timespec='seconds')}\n"
        f"{openpyxl_note}\n"
        "Contents:\n"
        f"  {active_stem}_phasor.png/.pdf/.svg — phasor plot (active sample)\n"
        f"  {active_stem}_segmentation.png     — paint overlay (if present)\n"
        "  samples_summary.csv               — metadata for all loaded samples\n"
        "  clusters.csv                      — cluster stats for all samples\n"
        "  analysis.xlsx                     — one workbook for this session\n"
        "                                      (rewritten on each Export all)\n"
        "  session.json                      — settings, cursors, GMM, stats\n"
        "  samples/<name>/                   — per-sample outputs (cleared &\n"
        "                                      rewritten on each export):\n"
        "    <file>_maps.npz                 — quantitative arrays:\n"
        f"                                     {keys_line}\n"
        "    <file>_photons.png              — masked intensity (no colorbar)\n"
        "    <file>_photons_colorbar.png     — masked intensity + colorbar\n"
        "    <file>_brightfield.png          — all photons (no colorbar)\n"
        "    <file>_brightfield_colorbar.png — all photons + colorbar\n"
        "    <file>_tau_*_ns.png             — lifetime map (no colorbar)\n"
        "    <file>_tau_*_ns_colorbar.png    — lifetime map + colorbar\n"
        "    <file>_clusters.csv             — cluster table for this sample\n"
        "    <file>_mask_XX.png              — cursor or GMM masks\n",
        encoding="utf-8",
    )
    written.append(str(readme))

    return {"directory": str(out), "files": written, "n_samples": len(datasets)}
