"""Main application window."""
import csv
import os
import sys
import time

import numpy as np
from matplotlib.backends.backend_qtagg import NavigationToolbar2QT as NavigationToolbar
from PySide6 import QtCore, QtGui, QtWidgets
from PySide6.QtCore import Qt

from flim_phasors.analysis import fit_phasor_gmm, label_pixels_by_gmm, lifetimes_at_phasor
from flim_phasors.constants import (
    COMPARE_STYLE_MAP,
    CURSOR_SHAPES,
    FILTER_MODES,
    FLIM_FILE_FILTER,
    IMAGE_VIEW_ITEMS,
)
from flim_phasors.data import PhasorData
from flim_phasors.io import is_supported_flim_path, load_reference_signal
from flim_phasors.canvas.image import ImageCanvas
from flim_phasors.canvas.phasor import PhasorCanvas
from flim_phasors.utils import (
    categorical_name,
    categorical_rgb,
    dataset_display_label,
    dataset_short_label,
)

try:
    from sklearn.mixture import GaussianMixture
    HAVE_SKLEARN = True
except ImportError:
    HAVE_SKLEARN = False

from phasorpy.cursor import mask_from_circular_cursor, mask_from_elliptic_cursor, pseudo_color
from phasorpy.lifetime import phasor_to_apparent_lifetime, phasor_to_normal_lifetime


class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("FLIM Phasor Analyzer — CAM segmentation")
        self.resize(1500, 980)
        self.data = PhasorData()
        self.datasets = []        # multi-image mode: list of PhasorData
        self.active_idx = -1
        self.shared_ref_path = ""
        self.shared_ref_n_channels = 1
        self.shared_ref_channel = 0
        self.last_overlay = None
        self.cluster_stats = []
        self.mode = "cursor"
        self._label_signal_connected = False
        self._paint_timer = QtCore.QTimer(self)
        self._paint_timer.setSingleShot(True)
        self._paint_timer.setInterval(250)   # ms after interaction stops -> full recompute
        self._paint_timer.timeout.connect(self._deferred_full_compute)
        self._build_ui()
        QtGui.QShortcut(QtGui.QKeySequence(Qt.Key.Key_Delete), self, self.remove_cursor)
        QtGui.QShortcut(QtGui.QKeySequence(Qt.Key.Key_Backspace), self, self.remove_cursor)

    # ---- UI ----------------------------------------------------------------
    def _build_ui(self):
        central = QtWidgets.QWidget(); self.setCentralWidget(central)
        main = QtWidgets.QHBoxLayout(central)

        panel_scroll = QtWidgets.QScrollArea()
        panel_scroll.setWidgetResizable(True)
        panel_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        panel_scroll.setFrameShape(QtWidgets.QFrame.Shape.NoFrame)
        panel_inner = QtWidgets.QWidget()
        pl = QtWidgets.QVBoxLayout(panel_inner)
        pl.setAlignment(Qt.AlignmentFlag.AlignTop)
        pl.setSpacing(5)
        pl.setContentsMargins(4, 4, 4, 4)

        _small = "font-size: 10px;"
        _lbl_file = f"color: gray; {_small}"

        # ---- files (sample | reference) + log ----
        gb_io = QtWidgets.QGroupBox("1 · Files")
        io_main = QtWidgets.QVBoxLayout(gb_io)
        io_main.setSpacing(4)
        io_row = QtWidgets.QHBoxLayout()
        io_row.setSpacing(6)

        sample_col = QtWidgets.QWidget()
        sl = QtWidgets.QVBoxLayout(sample_col); sl.setContentsMargins(0, 0, 0, 0); sl.setSpacing(2)
        row_s = QtWidgets.QHBoxLayout(); row_s.setSpacing(4)
        btn_sample = QtWidgets.QPushButton("Sample…")
        btn_sample.setMinimumWidth(72)
        btn_sample.setToolTip(
            "Load one or more FLIM files (.ptu, .tif). "
            "In the file dialog, Ctrl+click or Shift+click to select multiple.")
        btn_sample.clicked.connect(self.choose_sample)
        # btn_demo = QtWidgets.QPushButton("Demo")
        # btn_demo.clicked.connect(self.load_demo)
        self.cb_channel = QtWidgets.QComboBox()
        self.cb_channel.addItem("0")
        self.cb_channel.setMinimumWidth(48)
        self.cb_channel.currentIndexChanged.connect(self.on_channel_change)
        row_s.addWidget(btn_sample, 1)
        ch_lbl = QtWidgets.QLabel("Ch")
        ch_lbl.setSizePolicy(QtWidgets.QSizePolicy.Policy.Fixed, QtWidgets.QSizePolicy.Policy.Fixed)
        row_s.addWidget(ch_lbl)
        row_s.addWidget(self.cb_channel)
        # row_s.addWidget(btn_demo)
        sl.addLayout(row_s)
        self.lbl_sample = QtWidgets.QLabel("(no sample)")
        self.lbl_sample.setStyleSheet(_lbl_file)
        self.lbl_sample.setWordWrap(True)
        sl.addWidget(self.lbl_sample)

        ref_col = QtWidgets.QWidget()
        rl = QtWidgets.QVBoxLayout(ref_col); rl.setContentsMargins(0, 0, 0, 0); rl.setSpacing(2)
        self.chk_shared_ref = QtWidgets.QCheckBox("Shared ref")
        self.chk_shared_ref.setChecked(True)
        self.chk_shared_ref.setToolTip(
            "One reference calibrates all samples when checked.")
        self.chk_shared_ref.toggled.connect(self.on_shared_ref_toggle)
        rl.addWidget(self.chk_shared_ref)
        row_r = QtWidgets.QHBoxLayout(); row_r.setSpacing(4)
        btn_ref = QtWidgets.QPushButton("Reference…")
        btn_ref.setMinimumWidth(72)
        btn_ref.clicked.connect(self.choose_ref)
        self.cb_ref_channel = QtWidgets.QComboBox()
        self.cb_ref_channel.addItem("0")
        self.cb_ref_channel.setMinimumWidth(48)
        self.cb_ref_channel.setEnabled(False)
        self.cb_ref_channel.currentIndexChanged.connect(self.on_ref_channel_change)
        row_r.addWidget(btn_ref, 1)
        ref_ch_lbl = QtWidgets.QLabel("Ch")
        ref_ch_lbl.setSizePolicy(QtWidgets.QSizePolicy.Policy.Fixed, QtWidgets.QSizePolicy.Policy.Fixed)
        row_r.addWidget(ref_ch_lbl)
        row_r.addWidget(self.cb_ref_channel)
        rl.addLayout(row_r)
        self.lbl_ref = QtWidgets.QLabel("(none)")
        self.lbl_ref.setStyleSheet(_lbl_file)
        self.lbl_ref.setWordWrap(True)
        rl.addWidget(self.lbl_ref)

        io_row.addWidget(sample_col, 1)
        io_row.addWidget(ref_col, 1)
        io_main.addLayout(io_row)
        self.txt_log = QtWidgets.QPlainTextEdit()
        self.txt_log.setReadOnly(True)
        self.txt_log.setMaximumBlockCount(400)
        self.txt_log.setMinimumHeight(72)
        self.txt_log.setMaximumHeight(100)
        self.txt_log.setPlaceholderText("Activity log…")
        self.txt_log.setStyleSheet(
            f"font-family: Consolas, monospace; {_small} background: palette(base);")
        io_main.addWidget(self.txt_log)
        pl.addWidget(gb_io)

        # ---- calibration ----
        gb_proc = QtWidgets.QGroupBox("2 · Calibration")
        prg = QtWidgets.QGridLayout(gb_proc)
        prg.setHorizontalSpacing(6)
        prg.setVerticalSpacing(3)
        self.sp_harm = QtWidgets.QSpinBox(); self.sp_harm.setRange(1, 8); self.sp_harm.setValue(1)
        self.sp_freq = QtWidgets.QDoubleSpinBox(); self.sp_freq.setRange(1, 1000)
        self.sp_freq.setDecimals(3); self.sp_freq.setValue(80.0); self.sp_freq.setSuffix(" MHz")
        self.sp_reflt = QtWidgets.QDoubleSpinBox(); self.sp_reflt.setRange(0.0, 100.0)
        self.sp_reflt.setDecimals(3); self.sp_reflt.setValue(4.0); self.sp_reflt.setSuffix(" ns")
        self.sp_reflt.setToolTip(
            "Known lifetime of the reference dye (ns), e.g. ~4 for fluorescein. "
            "Wrong value shifts the phasor cloud vertically on the plot.")
        self.cb_filter = QtWidgets.QComboBox()
        self.cb_filter.addItems(list(FILTER_MODES))
        self.cb_filter.setCurrentText("median")
        self.cb_filter.currentTextChanged.connect(self.on_filter_change)
        self.chk_detect_harm = QtWidgets.QCheckBox("Harmonic mask")
        self.chk_detect_harm.setChecked(True)
        self.chk_detect_harm.setToolTip(
            "With Min N > 0, also mask harmonic-overtone pixels (phasorpy phasor_threshold).")
        self.sp_msize = QtWidgets.QSpinBox(); self.sp_msize.setRange(3, 11); self.sp_msize.setSingleStep(2); self.sp_msize.setValue(3)
        self.sp_mrep = QtWidgets.QSpinBox(); self.sp_mrep.setRange(1, 10); self.sp_mrep.setValue(1)
        self.sp_psigma = QtWidgets.QDoubleSpinBox(); self.sp_psigma.setRange(0.5, 6.0)
        self.sp_psigma.setSingleStep(0.5); self.sp_psigma.setValue(2.0)
        self.sp_plevels = QtWidgets.QSpinBox(); self.sp_plevels.setRange(1, 6); self.sp_plevels.setValue(1)
        self.sp_thr = QtWidgets.QSpinBox()
        self.sp_thr.setRange(0, 2_000_000_000)
        self.sp_thr.setSingleStep(100)
        self.sp_thr.setValue(0)
        self.sp_thr.setToolTip(
            "Minimum total photon count per pixel (sum of the TCSPC histogram). "
            "Pixels below this count are excluded from the phasor plot and segmentation only; "
            "the intensity image always shows all photons. 0 = off.")
        self.lbl_photon_range = QtWidgets.QLabel("(apply for photon range)")
        self.lbl_photon_range.setStyleSheet(_lbl_file)

        for sp in (self.sp_harm, self.sp_msize, self.sp_mrep, self.sp_plevels, self.sp_thr):
            sp.setMaximumWidth(72)
        for sp in (self.sp_freq, self.sp_reflt, self.sp_psigma):
            sp.setMaximumWidth(96)

        prg.addWidget(QtWidgets.QLabel("Harm."), 0, 0)
        prg.addWidget(self.sp_harm, 0, 1)
        prg.addWidget(QtWidgets.QLabel("Laser"), 0, 2)
        prg.addWidget(self.sp_freq, 0, 3)
        prg.addWidget(QtWidgets.QLabel("Ref τ"), 1, 0)
        prg.addWidget(self.sp_reflt, 1, 1)
        prg.addWidget(QtWidgets.QLabel("Filter"), 1, 2)
        prg.addWidget(self.cb_filter, 1, 3)
        self.lbl_msize = QtWidgets.QLabel("Kernel")
        prg.addWidget(self.lbl_msize, 2, 0)
        prg.addWidget(self.sp_msize, 2, 1)
        self.lbl_mrep = QtWidgets.QLabel("Repeat")
        prg.addWidget(self.lbl_mrep, 2, 2)
        prg.addWidget(self.sp_mrep, 2, 3)
        self.lbl_psigma = QtWidgets.QLabel("paw σ")
        prg.addWidget(self.lbl_psigma, 3, 0)
        prg.addWidget(self.sp_psigma, 3, 1)
        self.lbl_plevels = QtWidgets.QLabel("paw lvl")
        prg.addWidget(self.lbl_plevels, 3, 2)
        prg.addWidget(self.sp_plevels, 3, 3)
        self.row_msize = (self.lbl_msize, self.sp_msize, self.lbl_mrep, self.sp_mrep)
        self.row_psigma = (self.lbl_psigma, self.sp_psigma, self.lbl_plevels, self.sp_plevels)
        prg.addWidget(QtWidgets.QLabel("Min N"), 4, 0)
        prg.addWidget(self.sp_thr, 4, 1)
        prg.addWidget(self.chk_detect_harm, 4, 2, 1, 2)
        prg.addWidget(self.lbl_photon_range, 5, 0, 1, 4)
        btn_apply = QtWidgets.QPushButton("Apply"); btn_apply.clicked.connect(self.apply_processing)
        prg.addWidget(btn_apply, 6, 0, 1, 4)
        pl.addWidget(gb_proc)
        self.on_filter_change("median")

        # ---- multi-sample overlay ----
        gb_multi = QtWidgets.QGroupBox("3 · Multi-sample")
        mbl = QtWidgets.QVBoxLayout(gb_multi)
        mbl.setSpacing(3)
        self.chk_multi = QtWidgets.QCheckBox("Multi-image")
        self.chk_multi.toggled.connect(self.on_multi_toggle)
        mbl.addWidget(self.chk_multi)
        row_img = QtWidgets.QHBoxLayout()
        self.cb_image = QtWidgets.QComboBox()
        self.cb_image.setToolTip("Active sample for segmentation")
        self.cb_image.currentIndexChanged.connect(self.on_image_combo_change)
        btn_rmimg = QtWidgets.QPushButton("−"); btn_rmimg.setFixedWidth(28)
        btn_rmimg.setToolTip("Remove active sample from list")
        btn_rmimg.clicked.connect(self.remove_image)
        row_img.addWidget(self.cb_image, 1); row_img.addWidget(btn_rmimg)
        mbl.addLayout(row_img)
        row_grp = QtWidgets.QHBoxLayout()
        row_grp.addWidget(QtWidgets.QLabel("Group"))
        self.edit_group = QtWidgets.QLineEdit()
        self.edit_group.setPlaceholderText("e.g. Tumor, Control")
        self.edit_group.setToolTip(
            "Label for the active sample. Shown in the list and phasor overlay legend.")
        self.edit_group.editingFinished.connect(self._apply_group_from_field)
        btn_grp = QtWidgets.QPushButton("Apply")
        btn_grp.setFixedWidth(52)
        btn_grp.clicked.connect(self._apply_group_from_field)
        row_grp.addWidget(self.edit_group, 1)
        row_grp.addWidget(btn_grp)
        mbl.addLayout(row_grp)
        self.chk_compare = QtWidgets.QCheckBox("Phasor overlay")
        self.chk_compare.toggled.connect(self._on_compare_ui_changed)
        mbl.addWidget(self.chk_compare)
        row_cmp = QtWidgets.QHBoxLayout()
        self.cb_compare_style = QtWidgets.QComboBox()
        self.cb_compare_style.addItems(list(COMPARE_STYLE_MAP.keys()))
        self.cb_compare_style.currentIndexChanged.connect(self._on_compare_ui_changed)
        btn_cmp_all = QtWidgets.QPushButton("All"); btn_cmp_all.setFixedWidth(32)
        btn_cmp_all.clicked.connect(self._compare_select_all)
        btn_cmp_none = QtWidgets.QPushButton("None"); btn_cmp_none.setFixedWidth(36)
        btn_cmp_none.clicked.connect(self._compare_select_none)
        row_cmp.addWidget(self.cb_compare_style, 1)
        row_cmp.addWidget(btn_cmp_all)
        row_cmp.addWidget(btn_cmp_none)
        mbl.addLayout(row_cmp)
        self.table_compare = QtWidgets.QTableWidget(0, 5)
        self.table_compare.setHorizontalHeaderLabels(["Show", "#", "Group", "Sample", "Status"])
        self.table_compare.verticalHeader().setVisible(False)
        self.table_compare.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows)
        self.table_compare.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.SingleSelection)
        self.table_compare.setEditTriggers(
            QtWidgets.QAbstractItemView.EditTrigger.DoubleClicked
            | QtWidgets.QAbstractItemView.EditTrigger.EditKeyPressed)
        self.table_compare.setMinimumHeight(96)
        self.table_compare.setMaximumHeight(130)
        self.table_compare.setToolTip(
            "Group: name samples (e.g. condition). Tick Show to include on the multi-phasor plot.")
        hdr = self.table_compare.horizontalHeader()
        hdr.setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(2, QtWidgets.QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(3, QtWidgets.QHeaderView.ResizeMode.Stretch)
        hdr.setSectionResizeMode(4, QtWidgets.QHeaderView.ResizeMode.ResizeToContents)
        self.table_compare.cellChanged.connect(self._on_compare_table_changed)
        mbl.addWidget(self.table_compare, 1)
        self._compare_sel_buttons = (btn_cmp_all, btn_cmp_none)
        self._multi_detail_widgets = (
            self.cb_image, btn_rmimg, self.edit_group, btn_grp,
            self.chk_compare, self.cb_compare_style,
            self.table_compare, btn_cmp_all, btn_cmp_none,
        )
        self._set_multi_detail_enabled(False)
        self._set_compare_controls_enabled(False)
        pl.addWidget(gb_multi)
        self.gb_multi = gb_multi

        # ---- mode ----
        gb_mode = QtWidgets.QGroupBox("4 · Segmentation")
        ml = QtWidgets.QVBoxLayout(gb_mode)
        ml.setSpacing(3)
        mode_row = QtWidgets.QHBoxLayout()
        self.rb_cursor = QtWidgets.QRadioButton("Cursors")
        self.rb_gmm = QtWidgets.QRadioButton("GMM")
        self.rb_cursor.setChecked(True); self.rb_cursor.toggled.connect(self.on_mode_change)
        mode_row.addWidget(self.rb_cursor); mode_row.addWidget(self.rb_gmm)
        ml.addLayout(mode_row)

        self.cursor_box = QtWidgets.QWidget()
        cbl = QtWidgets.QGridLayout(self.cursor_box); cbl.setContentsMargins(0, 0, 0, 0)
        b_add = QtWidgets.QPushButton("+"); b_add.setFixedWidth(28); b_add.clicked.connect(self.add_cursor)
        b_del = QtWidgets.QPushButton("Del"); b_del.clicked.connect(self.remove_cursor)
        b_clr = QtWidgets.QPushButton("Clear"); b_clr.clicked.connect(self.clear_cursors)
        cbl.addWidget(b_add, 0, 0); cbl.addWidget(b_del, 0, 1); cbl.addWidget(b_clr, 0, 2)
        row_cur = QtWidgets.QHBoxLayout()
        row_cur.addWidget(QtWidgets.QLabel("Move"))
        self.cb_active_cursor = QtWidgets.QComboBox()
        self.cb_active_cursor.setToolTip(
            "Choose which circle to move or resize. Drag anywhere on the phasor plot.")
        self.cb_active_cursor.currentIndexChanged.connect(self.on_active_cursor_change)
        row_cur.addWidget(self.cb_active_cursor, 1)
        cbl.addLayout(row_cur, 1, 0, 1, 3)
        row_shape = QtWidgets.QHBoxLayout()
        row_shape.addWidget(QtWidgets.QLabel("Shape"))
        self.cb_cursor_shape = QtWidgets.QComboBox()
        self.cb_cursor_shape.addItems(list(CURSOR_SHAPES))
        self.cb_cursor_shape.setToolTip("Circle or ellipse ROI on the phasor plot (phasorpy cursors).")
        row_shape.addWidget(self.cb_cursor_shape, 1)
        row_shape.addWidget(QtWidgets.QLabel("Aspect"))
        self.sld_aspect = QtWidgets.QSlider(Qt.Orientation.Horizontal)
        self.sld_aspect.setRange(30, 100)
        self.sld_aspect.setValue(65)
        self.sld_aspect.setToolTip("Ellipse minor/major axis ratio (%).")
        self.lbl_aspect = QtWidgets.QLabel("65%")
        self.sld_aspect.valueChanged.connect(
            lambda v: self.lbl_aspect.setText(f"{v}%"))
        row_shape.addWidget(self.sld_aspect, 1)
        row_shape.addWidget(self.lbl_aspect)
        cbl.addLayout(row_shape, 2, 0, 1, 3)
        self.sld_radius = QtWidgets.QSlider(Qt.Orientation.Horizontal); self.sld_radius.setRange(5, 400); self.sld_radius.setValue(50)
        self.sld_radius.valueChanged.connect(self.on_radius_slider)
        self.lbl_radius = QtWidgets.QLabel("r=0.05")
        cbl.addWidget(self.sld_radius, 3, 0, 1, 2); cbl.addWidget(self.lbl_radius, 3, 2)
        ml.addWidget(self.cursor_box)

        self.gmm_box = QtWidgets.QWidget()
        gbl = QtWidgets.QGridLayout(self.gmm_box); gbl.setContentsMargins(0, 0, 0, 0)
        gbl.addWidget(QtWidgets.QLabel("k"), 0, 0)
        self.edit_ncomp = QtWidgets.QLineEdit("3")
        self.edit_ncomp.setValidator(QtGui.QIntValidator(1, 12, self))
        self.edit_ncomp.setFixedWidth(32)
        self.edit_ncomp.setToolTip("Number of GMM components (1–12).")
        gbl.addWidget(self.edit_ncomp, 0, 1)
        gbl.addWidget(QtWidgets.QLabel("Cov"), 0, 2)
        self.cb_cov = QtWidgets.QComboBox(); self.cb_cov.addItems(["full", "tied", "diag", "spherical"])
        gbl.addWidget(self.cb_cov, 0, 3)
        self.chk_bic = QtWidgets.QCheckBox("BIC auto-k"); gbl.addWidget(self.chk_bic, 1, 0, 1, 2)
        gbl.addWidget(QtWidgets.QLabel("σ"), 2, 0)
        self.edit_gmm_sigma = QtWidgets.QLineEdit("2.0")
        self.edit_gmm_sigma.setValidator(QtGui.QDoubleValidator(0.5, 6.0, 2, self))
        self.edit_gmm_sigma.setFixedWidth(40)
        self.edit_gmm_sigma.setToolTip(
            "Ellipse scale for phasorpy phasor_cluster_gmm (95% contour at 2.0).")
        gbl.addWidget(self.edit_gmm_sigma, 2, 1)
        b_fit = QtWidgets.QPushButton("Fit GMM"); b_fit.clicked.connect(self.fit_gmm)
        b_clr_gmm = QtWidgets.QPushButton("Clear"); b_clr_gmm.clicked.connect(self.clear_gmm)
        gbl.addWidget(b_fit, 2, 2); gbl.addWidget(b_clr_gmm, 2, 3)
        self.gmm_box.setVisible(False); ml.addWidget(self.gmm_box)
        pl.addWidget(gb_mode)

        # ---- actions ----
        gb_act = QtWidgets.QGroupBox("5 · Export")
        al = QtWidgets.QVBoxLayout(gb_act)
        al.setSpacing(3)
        row_paint = QtWidgets.QHBoxLayout()
        b_paint = QtWidgets.QPushButton("Paint"); b_paint.clicked.connect(self.compute_and_paint)
        self.chk_live = QtWidgets.QCheckBox("Live"); self.chk_live.setChecked(True)
        self.chk_overlay = QtWidgets.QCheckBox("Overlay"); self.chk_overlay.setChecked(True)
        self.chk_overlay.stateChanged.connect(self.refresh_image)
        row_paint.addWidget(b_paint, 1); row_paint.addWidget(self.chk_live); row_paint.addWidget(self.chk_overlay)
        al.addLayout(row_paint)
        row = QtWidgets.QHBoxLayout()
        b1 = QtWidgets.QPushButton("PNG seg"); b1.clicked.connect(self.save_overlay)
        b2 = QtWidgets.QPushButton("PNG phasor"); b2.clicked.connect(self.save_phasor)
        row.addWidget(b1); row.addWidget(b2); al.addLayout(row)
        rowx = QtWidgets.QHBoxLayout()
        b3 = QtWidgets.QPushButton("CSV"); b3.clicked.connect(self.export_csv)
        b4 = QtWidgets.QPushButton("Excel"); b4.clicked.connect(self.export_xlsx)
        rowx.addWidget(b3); rowx.addWidget(b4); al.addLayout(rowx)
        pl.addWidget(gb_act)
        panel_scroll.setWidget(panel_inner)
        panel_wrap = QtWidgets.QWidget()
        panel_wrap.setFixedWidth(420)
        pwl = QtWidgets.QVBoxLayout(panel_wrap)
        pwl.setContentsMargins(0, 0, 0, 0)
        pwl.addWidget(panel_scroll)

        # ---- plots ----
        plots = QtWidgets.QSplitter(Qt.Orientation.Horizontal)
        lw = QtWidgets.QWidget(); lv = QtWidgets.QVBoxLayout(lw)
        self.phasor = PhasorCanvas(self)
        self.phasor.cursorChanged.connect(self.on_cursor_changed)
        self.phasor.cursorMoving.connect(self.on_cursor_moving)
        self.phasor.cursorSelectionChanged.connect(self._on_phasor_cursor_selected)
        self.phasor.mpl_connect("key_press_event", self._on_phasor_key_press)
        self.phasor.phasorClicked.connect(self._on_phasor_click)
        self.phasor.setFocusPolicy(Qt.FocusPolicy.StrongFocus)
        lv.addWidget(NavigationToolbar(self.phasor, self)); lv.addWidget(self.phasor)
        rw = QtWidgets.QWidget(); rv = QtWidgets.QVBoxLayout(rw)
        row_img_view = QtWidgets.QHBoxLayout()
        row_img_view.addWidget(QtWidgets.QLabel("View"))
        self.cb_image_view = QtWidgets.QComboBox()
        self.cb_image_view.addItems(list(IMAGE_VIEW_ITEMS))
        self.cb_image_view.setToolTip(
            "Pixel maps from the current Apply settings (same mask as the phasor plot).")
        self.cb_image_view.currentIndexChanged.connect(self.refresh_image)
        row_img_view.addWidget(self.cb_image_view, 1)
        rv.addLayout(row_img_view)
        self.image = ImageCanvas(self)
        rv.addWidget(NavigationToolbar(self.image, self)); rv.addWidget(self.image)
        plots.addWidget(lw); plots.addWidget(rw); plots.setSizes([700, 700])

        self.table = QtWidgets.QTableWidget(0, 10)
        self.table.setHorizontalHeaderLabels(
            ["#", "Color", "Label (what you see)", "g", "s",
             "tau_phi (ns)", "tau_mod (ns)", "tau_normal (ns)",
             "Pixels", "Area %"])
        self.table.horizontalHeader().setSectionResizeMode(2, QtWidgets.QHeaderView.ResizeMode.Stretch)
        self.table.setMaximumHeight(220)

        rightside = QtWidgets.QSplitter(Qt.Orientation.Vertical)
        rightside.addWidget(plots); rightside.addWidget(self.table); rightside.setSizes([720, 220])
        main.addWidget(rightside, 1)
        main.addWidget(panel_wrap)

        self.status = self.statusBar()
        self._log("Ready — load a sample and optional reference.")

    def _log(self, message, update_status=True):
        """Append a timestamped line to the Files log and optionally the status bar."""
        line = f"[{time.strftime('%H:%M:%S')}] {message}"
        if hasattr(self, "txt_log"):
            self.txt_log.blockSignals(True)
            self.txt_log.appendPlainText(line)
            self.txt_log.blockSignals(False)
            sb = self.txt_log.verticalScrollBar()
            sb.setValue(sb.maximum())
        if update_status and hasattr(self, "status"):
            self.status.showMessage(message)

    def _form_row(self, form, label, widget):
        """Add a labelled row and return a (label_widget, field_widget) tuple for show/hide."""
        lbl = QtWidgets.QLabel(label)
        form.addRow(lbl, widget)
        return (lbl, widget)

    def _run_busy(self, message: str, fn):
        """Run a blocking call with a modal progress dialog; return (result, seconds)."""
        self._log(message, update_status=False)
        dlg = QtWidgets.QProgressDialog(message, None, 0, 0, self)
        dlg.setWindowModality(Qt.WindowModality.WindowModal)
        dlg.setMinimumDuration(0)
        dlg.setCancelButton(None)
        dlg.show()
        QtWidgets.QApplication.processEvents()
        t0 = time.perf_counter()
        try:
            return fn(), time.perf_counter() - t0
        finally:
            dlg.close()

    @staticmethod
    def _fmt_elapsed(seconds: float) -> str:
        if seconds < 1.0:
            return f"{seconds * 1000:.0f} ms"
        return f"{seconds:.2f} s"

    # ---- show/hide filter params ------------------------------------------
    def on_filter_change(self, mode):
        is_kernel = mode in ("median", "gaussian", "signal median", "signal gaussian")
        is_paw = mode == "pawflim"
        for w in self.row_msize:
            w.setVisible(is_kernel)
        for w in self.row_psigma:
            w.setVisible(is_paw)

    # ---- mode switching ----------------------------------------------------
    def on_mode_change(self):
        self.mode = "cursor" if self.rb_cursor.isChecked() else "gmm"
        self.cursor_box.setVisible(self.mode == "cursor")
        self.gmm_box.setVisible(self.mode == "gmm")
        if self.mode == "cursor":
            self.clear_gmm()
        else:
            self.phasor.clear_cursors()
            self.last_overlay = None
            self.cluster_stats = []
            self._fill_table()

    # ---- file actions ------------------------------------------------------
    def choose_sample(self):
        paths, _ = QtWidgets.QFileDialog.getOpenFileNames(
            self, "Choose sample FLIM file(s)", "", FLIM_FILE_FILTER)
        if not paths:
            return
        supported = []
        skipped = 0
        for path in paths:
            if is_supported_flim_path(path):
                supported.append(path)
            else:
                skipped += 1
                self._log(f"Skipped unsupported file: {os.path.basename(path)}")
        if not supported:
            QtWidgets.QMessageBox.warning(
                self, "Unsupported file",
                "Use PicoQuant .ptu or Imspector .tif / .tiff FLIM stacks.",
            )
            return
        if skipped:
            self._log(f"Skipped {skipped} unsupported file(s).")
        if not self._prepare_sample_load(supported):
            return
        self._load_sample_paths(supported)

    def _prepare_sample_load(self, paths):
        """Confirm how new paths merge with the current session. Returns False if cancelled."""
        has_current = self.data.signal_full is not None or self.data.is_synthetic
        batch = len(paths) > 1

        if batch:
            self.chk_multi.setChecked(True)
            if has_current:
                mbox = QtWidgets.QMessageBox(self)
                mbox.setWindowTitle("Load multiple samples")
                mbox.setText(f"Load {len(paths)} file(s).")
                mbox.setInformativeText(
                    "Replace all currently loaded samples, or add the new files to the list?")
                btn_replace = mbox.addButton(
                    "Replace all", QtWidgets.QMessageBox.ButtonRole.AcceptRole)
                btn_add = mbox.addButton(
                    "Add to list", QtWidgets.QMessageBox.ButtonRole.ActionRole)
                mbox.addButton(
                    QtWidgets.QMessageBox.StandardButton.Cancel)
                mbox.setDefaultButton(btn_add)
                mbox.exec()
                clicked = mbox.clickedButton()
                if clicked is None or clicked == mbox.button(QtWidgets.QMessageBox.StandardButton.Cancel):
                    return False
                if clicked == btn_replace:
                    self.datasets.clear()
                    self.active_idx = -1
                elif self.data not in self.datasets:
                    self.datasets.append(self.data)
            return True

        if has_current and not self.chk_multi.isChecked():
            btn = QtWidgets.QMessageBox.question(
                self, "Multi-image mode",
                "Keep the current sample and load another? This turns on multi-image mode "
                "so you can use the overlay table and shared reference.",
                QtWidgets.QMessageBox.StandardButton.Yes
                | QtWidgets.QMessageBox.StandardButton.No
                | QtWidgets.QMessageBox.StandardButton.Cancel,
            )
            if btn == QtWidgets.QMessageBox.StandardButton.Cancel:
                return False
            if btn == QtWidgets.QMessageBox.StandardButton.Yes:
                self.chk_multi.setChecked(True)
                if self.data not in self.datasets:
                    self.datasets.append(self.data)
            elif btn == QtWidgets.QMessageBox.StandardButton.No:
                self.datasets.clear()
                self.active_idx = -1
        elif has_current and self.chk_multi.isChecked() and self.data not in self.datasets:
            self.datasets.append(self.data)
        return True

    def _load_sample_paths(self, paths):
        """Decode one or more FLIM files and run phasor processing."""
        n = len(paths)
        loaded = []
        t_decode = 0.0
        for i, path in enumerate(paths):
            d = PhasorData()
            try:
                (shape, nch), elapsed = self._run_busy(
                    f"Decoding {i + 1}/{n}: {os.path.basename(path)}…",
                    lambda p=path, ds=d: ds.load_sample(p),
                )
                t_decode += elapsed
                loaded.append((d, path, shape, nch))
            except Exception as e:
                self._log(f"Load error ({os.path.basename(path)}): {e}")
                QtWidgets.QMessageBox.critical(
                    self, "Load error", f"{os.path.basename(path)}:\n{e}")
                if not loaded:
                    return
                break

        for d, path, shape, nch in loaded:
            self._activate_new_dataset(d)
            self._log(
                f"Loaded {os.path.basename(path)} — {shape[1]}×{shape[0]}, {nch} ch, "
                f"{d.frequency:.2f} MHz")

        if self.chk_multi.isChecked() and self.datasets:
            t0 = time.perf_counter()
            for d in self.datasets:
                if d.signal_full is not None:
                    self._run_processing_on_dataset(d, use_ui_settings=False)
            if 0 <= self.active_idx < len(self.datasets):
                self.data = self.datasets[self.active_idx]
            self._restore_ui_for_active()
            self._refresh_compare_list()
            self._update_phasor_display()
            self.chk_overlay.blockSignals(True)
            self.chk_overlay.setChecked(False)
            self.chk_overlay.blockSignals(False)
            self.refresh_image()
            t_proc = time.perf_counter() - t0
            self._log(
                f"{len(loaded)} sample(s) ready — decode {self._fmt_elapsed(t_decode)}, "
                f"phasor {self._fmt_elapsed(t_proc)} (multi-image).")
        else:
            _, t_proc = self._run_busy("Computing phasor…", self.apply_processing)
            d, path, shape, nch = loaded[-1]
            self._log(
                f"Loaded {os.path.basename(path)} — decode {self._fmt_elapsed(t_decode)}, "
                f"phasor {self._fmt_elapsed(t_proc)}")

    def _effective_ref_path(self, d=None):
        d = d or self.data
        if self.chk_shared_ref.isChecked() and self.shared_ref_path:
            return self.shared_ref_path
        return d.ref_path or None

    def _ref_channel_for_dataset(self, d):
        if self.chk_shared_ref.isChecked() and self.shared_ref_path:
            return min(self.shared_ref_channel, max(0, self.shared_ref_n_channels - 1))
        return min(d.ref_channel, max(0, d.ref_n_channels - 1)) if d.ref_path else 0

    def _propagate_shared_reference(self):
        """Copy shared reference settings onto every loaded sample when shared mode is on."""
        if not self.chk_shared_ref.isChecked():
            return
        path = self.shared_ref_path
        if not path:
            return
        for d in self._all_datasets():
            d.ref_path = path
            d.ref_n_channels = self.shared_ref_n_channels
            d.ref_channel = min(self.shared_ref_channel, d.ref_n_channels - 1)

    def on_shared_ref_toggle(self, checked):
        if checked and self.shared_ref_path:
            self._propagate_shared_reference()
            if any(d.signal_full is not None or d.is_synthetic for d in self._all_datasets()):
                self._recompute_all_datasets()
        self._restore_ui_for_active()
        self._log(
            "Shared reference on — one reference file calibrates all samples."
            if checked else
            "Shared reference off — each sample can use its own reference (active sample).")

    def _all_datasets(self):
        seen = []
        for d in [self.data] + list(self.datasets):
            if d is not None and d not in seen:
                seen.append(d)
        return seen

    def choose_ref(self):
        path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, "Choose reference FLIM file", "", FLIM_FILE_FILTER)
        if not path:
            return
        if not is_supported_flim_path(path):
            QtWidgets.QMessageBox.warning(
                self, "Unsupported file",
                "Use a PicoQuant .ptu or Imspector .tif / .tiff FLIM stack.",
            )
            return
        try:
            rsig, t_ref = self._run_busy(
                f"Decoding reference ({os.path.basename(path)})…",
                lambda: load_reference_signal(path))
        except Exception as e:
            self._log(f"Reference load error: {e}")
            QtWidgets.QMessageBox.critical(self, "Reference load error", str(e)); return
        ref_nch = int(rsig.sizes["C"]) if "C" in rsig.dims else 1
        self.shared_ref_path = path
        self.shared_ref_n_channels = ref_nch
        self.shared_ref_channel = min(self.shared_ref_channel, ref_nch - 1)
        if self.data.signal_full is not None:
            self.shared_ref_channel = min(self.data.channel, ref_nch - 1)
        self.lbl_ref.setText(os.path.basename(path))
        self._update_ref_channel_combo()
        if self.chk_shared_ref.isChecked():
            self._propagate_shared_reference()
            if any(d.signal_full is not None or d.is_synthetic for d in self._all_datasets()):
                self._recompute_all_datasets()
                self._log(
                    f"Shared reference ch {self.shared_ref_channel} ({self._fmt_elapsed(t_ref)}); "
                    "all loaded samples recalibrated.")
            else:
                self._log(
                    f"Shared reference — {ref_nch} channel(s) ({self._fmt_elapsed(t_ref)}). "
                    "Load samples, then Apply.")
        else:
            self.data.ref_path = path
            self.data.ref_n_channels = ref_nch
            self.data.ref_channel = min(self.data.channel, ref_nch - 1)
            if self.data.signal_full is not None or self.data.is_synthetic:
                self.apply_processing()
            self._log(
                f"Reference for active sample only, ch {self.data.ref_channel} "
                f"({self._fmt_elapsed(t_ref)}).")

    # def load_demo(self):
    #     """Synthetic CAM-like phasor demo (disabled in UI)."""
    #     d = PhasorData()
    #     d.load_synthetic()
    #     self._activate_new_dataset(d)
    #     self.lbl_ref.setText("(none — already calibrated)")
    #     if self.cb_filter.currentText() == "pawflim":
    #         self.cb_filter.setCurrentText("median")
    #     self.apply_processing()
    #     self._log(
    #         "Synthetic CAM-like demo loaded "
    #         "(background 0.4 ns / collagen 0.25 ns / vessels 2.0 ns).")

    # ---- multi-image management -------------------------------------------
    def _activate_new_dataset(self, d):
        """Make d the active dataset; append to the set if multi-image mode is on."""
        if self.chk_shared_ref.isChecked() and self.shared_ref_path:
            d.ref_path = self.shared_ref_path
            d.ref_n_channels = self.shared_ref_n_channels
            d.ref_channel = min(self.shared_ref_channel, d.ref_n_channels - 1)
        if self.chk_multi.isChecked():
            self.datasets.append(d)
            self.active_idx = len(self.datasets) - 1
        self.data = d
        self._restore_ui_for_active()
        if self.chk_multi.isChecked():
            self._refresh_image_combo()

    @staticmethod
    def _compact_filename(path, fallback="(none)", max_len=30):
        name = os.path.basename(path) if path else fallback
        if len(name) > max_len:
            return name[: max_len - 1] + "…"
        return name

    def _restore_ui_for_active(self):
        d = self.data
        self.lbl_sample.setText(self._compact_filename(d.sample_path, "<synthetic demo>"))
        if self.chk_shared_ref.isChecked():
            ref = self.shared_ref_path
        else:
            ref = d.ref_path
        self.lbl_ref.setText(self._compact_filename(ref) if ref else "(none)")
        nch = max(1, d.n_channels)
        self.cb_channel.blockSignals(True)
        self.cb_channel.clear()
        self.cb_channel.addItems([str(i) for i in range(nch)])
        self.cb_channel.setCurrentIndex(min(d.channel, nch - 1))
        self.cb_channel.blockSignals(False)
        self._update_ref_channel_combo()
        self.sp_freq.setValue(d.frequency)
        self.sp_harm.setValue(d.harmonic)
        if hasattr(self, "edit_group"):
            self.edit_group.blockSignals(True)
            self.edit_group.setText((d.group_name or "").strip())
            self.edit_group.blockSignals(False)

    def _update_ref_channel_combo(self):
        ref_path = self._effective_ref_path(self.data)
        has_ref = bool(ref_path)
        if self.chk_shared_ref.isChecked() and self.shared_ref_path:
            nch = max(1, self.shared_ref_n_channels)
            ch = self.shared_ref_channel
        else:
            nch = max(1, self.data.ref_n_channels)
            ch = self.data.ref_channel
        self.cb_ref_channel.blockSignals(True)
        self.cb_ref_channel.clear()
        self.cb_ref_channel.addItems([str(i) for i in range(nch)])
        if has_ref:
            self.cb_ref_channel.setCurrentIndex(min(ch, nch - 1))
        self.cb_ref_channel.setEnabled(has_ref)
        self.cb_ref_channel.blockSignals(False)

    def _set_multi_detail_enabled(self, enabled):
        for w in self._multi_detail_widgets:
            w.setEnabled(enabled)

    def _recompute_all_datasets(self):
        """Re-run phasor pipeline on every loaded sample (e.g. after shared reference change)."""
        targets = list(self.datasets) if self.chk_multi.isChecked() and self.datasets else []
        if not targets:
            self.apply_processing()
            return
        saved = self.data
        saved_idx = self.active_idx
        for d in targets:
            if d.signal_full is None and not d.is_synthetic:
                continue
            self._run_processing_on_dataset(d, use_ui_settings=False)
        if 0 <= saved_idx < len(self.datasets):
            self.data = self.datasets[saved_idx]
            self.active_idx = saved_idx
        else:
            self.data = saved
        self._restore_ui_for_active()
        self._refresh_compare_list()
        self._update_phasor_display()
        self.refresh_image()

    def _refresh_image_combo(self):
        self.cb_image.blockSignals(True)
        self.cb_image.clear()
        for i, d in enumerate(self.datasets):
            self.cb_image.addItem(f"{i + 1}: {dataset_display_label(d, i)}")
        if 0 <= self.active_idx < len(self.datasets):
            self.cb_image.setCurrentIndex(self.active_idx)
        self.cb_image.blockSignals(False)
        self._refresh_compare_list()

    def _apply_group_from_field(self):
        text = self.edit_group.text().strip() if hasattr(self, "edit_group") else ""
        if self.chk_multi.isChecked() and 0 <= self.active_idx < len(self.datasets):
            self.datasets[self.active_idx].group_name = text
        else:
            self.data.group_name = text
        self._refresh_image_combo()
        if self.data.real_cal is not None:
            self._update_phasor_display()
        if text:
            self._log(f"Group set to “{text}” for active sample.")

    def _compare_show_item(self, row):
        return self.table_compare.item(row, 0)

    def _compare_dataset_index(self, row):
        it = self._compare_show_item(row)
        if it is None:
            return -1
        return int(it.data(Qt.ItemDataRole.UserRole))

    def _refresh_compare_list(self):
        checked = {}
        for row in range(self.table_compare.rowCount()):
            idx = self._compare_dataset_index(row)
            it = self._compare_show_item(row)
            if idx >= 0 and it is not None:
                checked[idx] = it.checkState() == Qt.CheckState.Checked

        self.table_compare.blockSignals(True)
        self.table_compare.setRowCount(len(self.datasets))
        for i, d in enumerate(self.datasets):
            ready = d.real_cal is not None
            show = QtWidgets.QTableWidgetItem()
            show.setData(Qt.ItemDataRole.UserRole, i)
            show.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            if ready:
                show.setFlags(Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled)
                show.setCheckState(
                    Qt.CheckState.Checked if checked.get(i, True) else Qt.CheckState.Unchecked)
            else:
                show.setFlags(Qt.ItemFlag.ItemIsEnabled)
                show.setCheckState(Qt.CheckState.Unchecked)
                show.setToolTip("Run Apply on this image first")

            num = self._ro(str(i + 1))
            num.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            group = self._editable_group((d.group_name or "").strip())
            sample = self._ro(dataset_short_label(d, i))
            if i == self.active_idx:
                sample.setFont(QtGui.QFont(sample.font().family(), sample.font().pointSize(),
                                           QtGui.QFont.Weight.Bold))
                sample.setToolTip("Active image (dropdown) — used for segmentation")
            status = self._ro("ready" if ready else "not computed")
            if not ready:
                status.setForeground(QtGui.QBrush(Qt.GlobalColor.gray))

            self.table_compare.setItem(i, 0, show)
            self.table_compare.setItem(i, 1, num)
            self.table_compare.setItem(i, 2, group)
            self.table_compare.setItem(i, 3, sample)
            self.table_compare.setItem(i, 4, status)
        self.table_compare.blockSignals(False)
        self._set_compare_controls_enabled(
            self.chk_multi.isChecked() and len(self.datasets) >= 2)

    def _set_compare_controls_enabled(self, compare_available):
        multi_on = self.chk_multi.isChecked() and len(self.datasets) >= 1
        cmp_on = compare_available and self.chk_compare.isChecked()
        self.chk_compare.setEnabled(compare_available)
        self.cb_compare_style.setEnabled(cmp_on)
        self.table_compare.setEnabled(multi_on)
        for btn in getattr(self, "_compare_sel_buttons", ()):
            btn.setEnabled(cmp_on)

    def _compare_set_all_checks(self, checked):
        state = Qt.CheckState.Checked if checked else Qt.CheckState.Unchecked
        self.table_compare.blockSignals(True)
        for row in range(self.table_compare.rowCount()):
            it = self._compare_show_item(row)
            if it is None or not (it.flags() & Qt.ItemFlag.ItemIsUserCheckable):
                continue
            it.setCheckState(state)
        self.table_compare.blockSignals(False)
        self._update_phasor_display()

    def _compare_select_all(self):
        self._compare_set_all_checks(True)

    def _compare_select_none(self):
        self._compare_set_all_checks(False)

    def _on_compare_table_changed(self, row, column):
        if column == 2:
            if 0 <= row < len(self.datasets):
                item = self.table_compare.item(row, 2)
                text = item.text().strip() if item else ""
                self.datasets[row].group_name = text
                if row == self.active_idx and hasattr(self, "edit_group"):
                    self.edit_group.blockSignals(True)
                    self.edit_group.setText(text)
                    self.edit_group.blockSignals(False)
                self._refresh_image_combo()
                self._update_phasor_display()
            return
        if column != 0:
            return
        self._on_compare_ui_changed()

    def _build_compare_layers(self):
        layers = []
        for row in range(self.table_compare.rowCount()):
            idx = self._compare_dataset_index(row)
            if idx < 0 or idx >= len(self.datasets):
                continue
            show = self._compare_show_item(row)
            d = self.datasets[idx]
            layers.append({
                "data": d,
                "label": dataset_display_label(d, idx),
                "color": categorical_rgb(idx),
                "visible": show is not None and show.checkState() == Qt.CheckState.Checked,
                "index": idx,
            })
        return layers

    def _compare_style_key(self):
        return COMPARE_STYLE_MAP.get(self.cb_compare_style.currentText(), "cloud")

    def _update_phasor_display(self, status_note=""):
        layers = self._build_compare_layers()
        visible = [L for L in layers if L.get("visible") and L["data"].real_cal is not None]
        compare_on = (
            self.chk_multi.isChecked()
            and self.chk_compare.isChecked()
            and len(visible) >= 1
        )
        self.phasor.update_display(
            self.data,
            compare_enabled=compare_on,
            compare_style=self._compare_style_key(),
            compare_layers=layers if compare_on else None,
        )
        if compare_on and len(visible) >= 2:
            freqs = {round(L["data"].work_frequency, 4) for L in visible}
            harms = {L["data"].harmonic for L in visible}
            if len(freqs) > 1 or len(harms) > 1:
                self._log(
                    "Compare overlay: working frequency or harmonic differs between images — "
                    "positions may not be directly comparable.")
            elif status_note:
                self._log(status_note)

    def _on_compare_ui_changed(self, *_args):
        self._set_compare_controls_enabled(
            self.chk_multi.isChecked() and len(self.datasets) >= 2)
        if self.data.real_cal is not None or self.chk_compare.isChecked():
            self._update_phasor_display()

    def on_multi_toggle(self, checked):
        self._set_multi_detail_enabled(checked)
        if checked:
            # adopt the currently loaded image as the first slot (non-destructive)
            has_image = self.data.signal_full is not None or self.data.is_synthetic
            if has_image and self.data not in self.datasets:
                self.datasets.append(self.data)
                self.active_idx = len(self.datasets) - 1
            self._refresh_image_combo()
            self._log(
                "Multi-image mode on — use the overlay table below calibration "
                "(tick Shared reference to calibrate all samples with one file).")
        else:
            self.chk_compare.blockSignals(True)
            self.chk_compare.setChecked(False)
            self.chk_compare.blockSignals(False)
            self._set_compare_controls_enabled(False)
            self._update_phasor_display()
            self._log("Multi-image mode off (current image stays active).")

    def on_image_combo_change(self, idx):
        if not (0 <= idx < len(self.datasets)):
            return
        self.active_idx = idx
        self.data = self.datasets[idx]
        self._restore_ui_for_active()
        self._refresh_compare_list()
        # redisplay this image's already-processed phasor without recomputing
        self._update_phasor_display()
        self.last_overlay = None
        self.cluster_stats = []
        if self.chk_live.isChecked() and self.mode == "cursor" and self.phasor.cursors \
                and self.data.real_cal is not None:
            self._compute_cursor()                  # repaint shared cursors on this image
        else:
            self._fill_table()
            self.chk_overlay.blockSignals(True); self.chk_overlay.setChecked(False); self.chk_overlay.blockSignals(False)
            self.refresh_image()
        self._log(f"Switched to image {idx + 1}: {dataset_display_label(self.data, idx)}")

    def remove_image(self):
        if not (0 <= self.active_idx < len(self.datasets)):
            return
        name = dataset_short_label(self.data, self.active_idx)
        self.datasets.pop(self.active_idx)
        self._log(f"Removed sample: {name}")
        if not self.datasets:
            self.active_idx = -1
            self._refresh_image_combo()
            return
        self.active_idx = min(self.active_idx, len(self.datasets) - 1)
        self._refresh_image_combo()
        self.on_image_combo_change(self.active_idx)

    def on_channel_change(self, idx):
        if self.data.signal_full is None and not self.data.is_synthetic:
            return
        self.data.channel = max(0, idx)
        self.apply_processing()

    def on_ref_channel_change(self, idx):
        if not self._effective_ref_path(self.data):
            return
        if self.chk_shared_ref.isChecked() and self.shared_ref_path:
            ch = max(0, min(idx, self.shared_ref_n_channels - 1))
            self.shared_ref_channel = ch
            self._propagate_shared_reference()
            if any(d.signal_full is not None or d.is_synthetic for d in self._all_datasets()):
                self._recompute_all_datasets()
        else:
            self.data.ref_channel = max(0, min(idx, self.data.ref_n_channels - 1))
            if self.data.signal_full is not None or self.data.is_synthetic:
                self.apply_processing()

    # ---- processing --------------------------------------------------------
    def _processing_params(self, d):
        mode = self.cb_filter.currentText()
        if mode == "pawflim" and d.is_synthetic:
            if d is self.data:
                QtWidgets.QMessageBox.information(
                    self, "pawFLIM",
                    "pawFLIM needs a real TCSPC signal (multi-harmonic) and isn't available "
                    "for the synthetic demo. Using 'median' instead.")
                self.cb_filter.setCurrentText("median")
            mode = "median"
        return {
            "ref_path": self._effective_ref_path(d),
            "ref_lifetime": self.sp_reflt.value(),
            "filter_mode": mode,
            "median_size": self.sp_msize.value(),
            "median_repeat": self.sp_mrep.value(),
            "paw_sigma": self.sp_psigma.value(),
            "paw_levels": self.sp_plevels.value(),
            "intensity_min": float(self.sp_thr.value()),
            "detect_harmonics": self.chk_detect_harm.isChecked(),
        }

    def _run_processing_on_dataset(self, d, *, use_ui_settings=False):
        if use_ui_settings:
            d.harmonic = self.sp_harm.value()
            d.frequency = self.sp_freq.value()
            d.channel = max(0, self.cb_channel.currentIndex())
        if self._effective_ref_path(d):
            d.ref_channel = self._ref_channel_for_dataset(d)
        d.apply_processing(**self._processing_params(d))

    def apply_processing(self):
        if self.data.signal_full is None and not self.data.is_synthetic:
            QtWidgets.QMessageBox.information(self, "No data", "Load a sample first.")
            return
        t0 = time.perf_counter()
        try:
            self._run_processing_on_dataset(self.data, use_ui_settings=True)
        except Exception as e:
            self._log(f"Processing error: {e}")
            QtWidgets.QMessageBox.critical(self, "Processing error", repr(e)); return
        elapsed = time.perf_counter() - t0
        self._refresh_compare_list()
        self._update_phasor_display()
        self.chk_overlay.blockSignals(True); self.chk_overlay.setChecked(False); self.chk_overlay.blockSignals(False)
        self.refresh_image()
        if self._effective_ref_path(self.data):
            ch = self._ref_channel_for_dataset(self.data)
            ref_note = f", shared ref ch {ch}" if self.chk_shared_ref.isChecked() else f", ref ch {ch}"
        else:
            ref_note = ""
        st = getattr(self.data, "_intensity_stats", {})
        if st:
            self.lbl_photon_range.setText(
                f"Image photon counts: {st['min']:.0f} – {st['max']:.0f} "
                f"(median {st['median']:.0f})")
            n_below = int(round(st["masked_pct"] * st.get("n_pixels", 0) / 100.0))
            int_msg = (f" | min photons ≥ {st['threshold']:.0f} "
                       f"({n_below} px removed)")
        else:
            int_msg = ""
        self._log(
            f"Phasor recomputed — sample ch {self.data.channel}{ref_note}, "
            f"filter={self.cb_filter.currentText()}, H={self.data.harmonic}{int_msg}  "
            f"({self._fmt_elapsed(elapsed)})")

    # ---- cursor actions ----------------------------------------------------
    def _refresh_active_cursor_combo(self, select_idx=None):
        if not hasattr(self, "cb_active_cursor"):
            return
        idx = self.phasor.selected if select_idx is None else select_idx
        self.cb_active_cursor.blockSignals(True)
        self.cb_active_cursor.clear()
        for i, c in enumerate(self.phasor.cursors):
            label = c.get("label") or categorical_name(i)
            self.cb_active_cursor.addItem(label)
        enabled = len(self.phasor.cursors) > 0
        self.cb_active_cursor.setEnabled(enabled)
        if enabled and 0 <= idx < len(self.phasor.cursors):
            self.cb_active_cursor.setCurrentIndex(idx)
        self.cb_active_cursor.blockSignals(False)

    def _on_phasor_cursor_selected(self, idx):
        self._refresh_active_cursor_combo(select_idx=idx)
        self._sync_radius_slider()

    def on_active_cursor_change(self, combo_idx):
        if self.mode != "cursor" or combo_idx < 0:
            return
        self.phasor.select_cursor(combo_idx, emit=False)
        self._sync_radius_slider()

    def add_cursor(self):
        r = self.sld_radius.value() * 0.001
        kind = "ellipse" if self.cb_cursor_shape.currentText() == "Ellipse" else "circle"
        aspect = self.sld_aspect.value() / 100.0
        self.phasor.add_cursor(
            radius=r, kind=kind,
            radius_minor=r * aspect if kind == "ellipse" else None)
        self._refresh_active_cursor_combo()

    def remove_cursor(self):
        if self.mode != "cursor":
            return
        if self.phasor.selected < 0:
            self._log("Choose a circle in Move, or click one on the phasor plot.")
            return
        self.phasor.remove_selected()
        self._refresh_active_cursor_combo()
        self._refresh_after_cursor_edit()
        self._log("Circle removed.")

    def clear_cursors(self):
        self.phasor.clear_cursors()
        self._refresh_active_cursor_combo()
        self._refresh_after_cursor_edit()
        self._log("All circles cleared.")

    def _refresh_after_cursor_edit(self):
        """Update segmentation overlay and table whenever circles are added/removed/moved."""
        self._paint_timer.stop()
        if self.mode != "cursor" or self.data.real_cal is None:
            return
        if self.phasor.cursors:
            self._compute_cursor()
        else:
            self.last_overlay = None
            self.cluster_stats = []
            self._fill_table()
            self.chk_overlay.blockSignals(True)
            self.chk_overlay.setChecked(False)
            self.chk_overlay.blockSignals(False)
            self.refresh_image()

    def _on_phasor_key_press(self, event):
        if self.mode != "cursor":
            return
        if event.key in ("delete", "backspace"):
            self.remove_cursor()

    def clear_gmm(self):
        self.phasor.clear_gmm()
        if hasattr(self, "gmm"):
            del self.gmm
        if hasattr(self, "_gmm_fit"):
            del self._gmm_fit
        self.last_overlay = None
        self.cluster_stats = []
        self._fill_table()
        self.refresh_image()
        self._log("GMM fit cleared.")

    def on_radius_slider(self, v):
        r = v * 0.001
        self.lbl_radius.setText(f"r={r:.2f}")
        self.phasor.set_selected_radius(r)

    def _sync_radius_slider(self):
        i = self.phasor.selected
        if 0 <= i < len(self.phasor.cursors):
            r = self.phasor.cursors[i]["radius"]
            self.sld_radius.blockSignals(True); self.sld_radius.setValue(int(round(r * 1000))); self.sld_radius.blockSignals(False)
            self.lbl_radius.setText(f"{r:.3f}")

    def _live_active(self):
        return self.chk_live.isChecked() and self.mode == "cursor" and self.data.real_cal is not None

    def on_cursor_moving(self):
        """Interactive drag/scroll/slider: repaint overlay only (fast), defer the rest."""
        self._sync_radius_slider()
        if not self._live_active() or not self.phasor.cursors:
            return
        masks, colors = self._cursor_masks_colors()
        if masks is None:
            return
        intensity = self._segmentation_intensity()
        overlay = pseudo_color(*[masks[k] for k in range(len(masks))],
                               intensity=intensity, colors=np.array(colors))
        self.last_overlay = np.clip(np.asarray(overlay), 0, 1)
        if not self.chk_overlay.isChecked():
            self.chk_overlay.blockSignals(True); self.chk_overlay.setChecked(True); self.chk_overlay.blockSignals(False)
        self.image.update_overlay(self.last_overlay, title=f"Segmentation ({self.mode})")
        self._paint_timer.start()   # full recompute (lifetimes + table) once interaction settles

    def on_cursor_changed(self):
        """Committed change: sync slider and refresh segmentation (always, not only live-paint)."""
        self._refresh_active_cursor_combo()
        self._sync_radius_slider()
        self._refresh_after_cursor_edit()

    def _deferred_full_compute(self):
        if self._live_active() and self.phasor.cursors:
            self._compute_cursor()

    def _gmm_k_max(self) -> int:
        text = self.edit_ncomp.text().strip() if hasattr(self, "edit_ncomp") else "3"
        try:
            return max(1, min(12, int(text)))
        except ValueError:
            return 3

    def _gmm_sigma(self) -> float:
        text = self.edit_gmm_sigma.text().strip() if hasattr(self, "edit_gmm_sigma") else "2.0"
        try:
            return max(0.5, min(6.0, float(text)))
        except ValueError:
            return 2.0

    # ---- GMM ---------------------------------------------------------------
    def fit_gmm(self):
        if not HAVE_SKLEARN:
            QtWidgets.QMessageBox.warning(self, "Missing dependency", "pip install scikit-learn"); return
        if not self.rb_gmm.isChecked():
            self.rb_gmm.setChecked(True)
        if self.data.real_cal is None:
            QtWidgets.QMessageBox.information(self, "GMM", "Load data and click Apply first."); return
        m = self.data.valid_mask()
        if m.sum() < 10:
            QtWidgets.QMessageBox.information(self, "GMM", "Not enough valid pixels."); return
        g, s = self.data.real_cal, self.data.imag_cal
        cov = self.cb_cov.currentText()
        sigma = self._gmm_sigma()
        try:
            if self.chk_bic.isChecked():
                X = np.column_stack([g[m], s[m]])
                best_n, best_bic = 1, np.inf
                for n in range(1, self._gmm_k_max() + 1):
                    gm = GaussianMixture(n, covariance_type=cov, random_state=0).fit(X)
                    b = gm.bic(X)
                    if b < best_bic:
                        best_bic, best_n = b, n
                n_clusters = best_n
                self._log(f"GMM auto-selected {best_n} components (BIC={best_bic:.0f}).")
            else:
                n_clusters = self._gmm_k_max()
            self._gmm_fit = fit_phasor_gmm(
                g, s, clusters=n_clusters, sigma=sigma, covariance_type=cov)
            self._log(f"phasorpy GMM: {n_clusters} cluster(s), σ={sigma:.1f}.")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "GMM fit failed", str(e)); return
        n = len(self._gmm_fit[0])
        colors = [categorical_rgb(k) for k in range(n)]
        self.phasor.show_gmm_ellipses(*self._gmm_fit, colors)
        self._compute_gmm()
        return

    def _on_phasor_click(self, g, s):
        if self.data.real_cal is None or self.data.work_frequency <= 0:
            return
        try:
            tp, tm, tn = lifetimes_at_phasor(g, s, self.data.work_frequency)
        except Exception as e:
            self._log(f"Phasor readout failed: {e}")
            return
        self._log(
            f"Phasor click ({g:.3f}, {s:.3f}) → τφ={tp:.3f} ns, τmod={tm:.3f} ns, τn={tn:.3f} ns")

    # ---- compute lifetimes + paint ----------------------------------------
    def compute_and_paint(self):
        if self.data.real_cal is None:
            self._log("Paint skipped — run Apply first.")
            return
        self._log(f"Paint ({self.mode})…")
        if self.mode == "cursor":
            self._compute_cursor()
        else:
            self._compute_gmm()
        self._log("Paint complete.")

    def _lifetimes(self, g, s):
        freq = self.data.work_frequency
        tp, tm = phasor_to_apparent_lifetime(g, s, freq)
        tn = phasor_to_normal_lifetime(g, s, freq)
        return float(tp), float(tm), float(tn)

    def _cursor_masks_colors(self):
        """Boolean masks (within valid pixels) and colors for the current circles."""
        cur = self.phasor.cursors
        if not cur or self.data.real_cal is None:
            return None, None
        g, s = self.data.real_cal, self.data.imag_cal
        valid = self.data.valid_mask()
        masks = []
        for c in cur:
            if c.get("kind") == "ellipse":
                mk = mask_from_elliptic_cursor(
                    g, s, [c["center_real"]], [c["center_imag"]],
                    radius=[c["radius"]],
                    radius_minor=[c.get("radius_minor", c["radius"] * 0.65)],
                    angle=[c.get("angle", 0.0)],
                )
            else:
                mk = mask_from_circular_cursor(
                    g, s, [c["center_real"]], [c["center_imag"]], radius=[c["radius"]])
            if mk.ndim == 3:
                mk = mk[0]
            masks.append(mk & valid)
        masks = np.stack(masks)
        return masks, [c["color"] for c in cur]

    def _compute_cursor(self):
        cur = self.phasor.cursors
        if not cur:
            QtWidgets.QMessageBox.information(self, "Cursors", "Add at least one circle."); return
        g, s = self.data.real_cal, self.data.imag_cal
        masks, colors = self._cursor_masks_colors()
        valid = self.data.valid_mask()
        total_valid = max(int(valid.sum()), 1)
        self.cluster_stats = []
        for k, c in enumerate(cur):
            mk = masks[k]; n = int(mk.sum())
            if n > 0:
                cg = float(np.nanmean(g[mk])); cs = float(np.nanmean(s[mk]))
                tp, tm, tn = self._lifetimes(cg, cs)
            else:
                cg = cs = tp = tm = tn = float("nan")
            self.cluster_stats.append(dict(idx=k + 1, color=c["color"], label=c["label"],
                                           tp=tp, tm=tm, tn=tn, g=cg, s=cs, n=n,
                                           area=100.0 * n / total_valid))
        self._paint(masks, colors); self._fill_table()

    def _compute_gmm(self):
        if not hasattr(self, "_gmm_fit"):
            QtWidgets.QMessageBox.information(self, "GMM", "Fit a GMM first."); return
        cr, ci, rm, ri, ang = self._gmm_fit
        g, s = self.data.real_cal, self.data.imag_cal
        valid = self.data.valid_mask()
        if valid.sum() < 10:
            QtWidgets.QMessageBox.information(self, "GMM", "Not enough valid pixels."); return
        n_comp = len(cr)
        labelmap = label_pixels_by_gmm(g, s, cr, ci, rm)
        masks = np.stack([(labelmap == k) & valid for k in range(n_comp)])
        colors = [categorical_rgb(k) for k in range(n_comp)]
        total_valid = max(int(valid.sum()), 1)
        self.cluster_stats = []
        for k in range(n_comp):
            cg, cs = float(cr[k]), float(ci[k])
            tp, tm, tn = self._lifetimes(cg, cs)
            n = int(masks[k].sum())
            self.cluster_stats.append(dict(idx=k + 1, color=colors[k], label=categorical_name(k),
                                           tp=tp, tm=tm, tn=tn, g=cg, s=cs, n=n,
                                           area=100.0 * n / total_valid))
        self._paint(masks, colors); self._fill_table()

    def _phasor_valid_mask(self):
        if self.data.real_cal is None:
            return None
        return self.data.valid_mask()

    def _mask_like_phasor(self, arr):
        """Apply the same finite-pixel mask used for the phasor histogram."""
        if arr is None:
            return None
        out = np.asarray(arr, dtype=float).copy()
        valid = self._phasor_valid_mask()
        if valid is not None and valid.shape == out.shape:
            out[~valid] = np.nan
        return out

    def _photon_image_filtered(self):
        """Photon counts with NaN where excluded from phasor (threshold + invalid phasor)."""
        base = self.data.mean_thr if self.data.mean_thr is not None else self.data.mean_raw
        return self._mask_like_phasor(base)

    def _segmentation_intensity(self):
        """Background for pseudo_color — photons masked like the phasor plot."""
        raw = self._photon_image_filtered()
        if raw is None:
            return np.zeros((2, 2), dtype=np.float64)
        out = np.zeros(np.shape(raw), dtype=np.float64)
        finite = np.isfinite(raw)
        out[finite] = np.nan_to_num(raw[finite])
        return out

    def _paint(self, masks, colors):
        intensity = self._segmentation_intensity()
        overlay = pseudo_color(*[masks[k] for k in range(len(masks))],
                               intensity=intensity, colors=np.array(colors))
        self.last_overlay = np.clip(np.asarray(overlay), 0, 1)
        self.chk_overlay.setChecked(True); self.refresh_image()

    def refresh_image(self):
        if self.chk_overlay.isChecked() and self.last_overlay is not None:
            self.image.show_overlay(self.last_overlay, title=f"Segmentation ({self.mode})")
        else:
            self._show_base_image()

    def _show_base_image(self):
        """Render the view selected above the image (masked like the phasor plot)."""
        choice = (
            self.cb_image_view.currentText()
            if hasattr(self, "cb_image_view")
            else IMAGE_VIEW_ITEMS[0]
        )
        if choice == IMAGE_VIEW_ITEMS[0]:
            disp = self._photon_image_filtered()
            if disp is not None:
                self.image.show_intensity(disp)
            return
        tau_sources = {
            IMAGE_VIEW_ITEMS[1]: (self.data.tau_phi, "τφ phase (ns)"),
            IMAGE_VIEW_ITEMS[2]: (self.data.tau_mod, "τmod (ns)"),
            IMAGE_VIEW_ITEMS[3]: (self.data.tau_normal, "τ normal (ns)"),
            IMAGE_VIEW_ITEMS[4]: (self.data.tau_search_phi, "τ search phase (ns)"),
            IMAGE_VIEW_ITEMS[5]: (self.data.tau_search_mod, "τ search mod (ns)"),
        }
        src = tau_sources.get(choice)
        if src is None:
            return
        tau_map, cbar_label = src
        disp = self._mask_like_phasor(tau_map)
        if disp is None:
            return
        vmin, vmax = self._robust_range(disp)
        self.image.show_map(disp, title=choice, cmap="turbo", label=cbar_label,
                            vmin=vmin, vmax=vmax)

    @staticmethod
    def _robust_range(arr):
        finite = np.asarray(arr)[np.isfinite(arr)]
        if finite.size == 0:
            return None, None
        return float(np.nanpercentile(finite, 2)), float(np.nanpercentile(finite, 98))

    # ---- results table -----------------------------------------------------
    def _fill_table(self):
        if getattr(self, "_label_signal_connected", False):
            self.table.itemChanged.disconnect(self._label_edited)
            self._label_signal_connected = False
        self.table.setRowCount(len(self.cluster_stats))
        for r, st in enumerate(self.cluster_stats):
            self.table.setItem(r, 0, self._ro(str(st["idx"])))
            sw = QtWidgets.QTableWidgetItem(); sw.setBackground(QtGui.QColor.fromRgbF(*st["color"]))
            sw.setFlags(Qt.ItemFlag.ItemIsEnabled); self.table.setItem(r, 1, sw)
            lab = QtWidgets.QTableWidgetItem(st["label"])
            lab.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsEditable | Qt.ItemFlag.ItemIsSelectable)
            self.table.setItem(r, 2, lab)
            self.table.setItem(r, 3, self._ro(f"{st['g']:.4f}"))
            self.table.setItem(r, 4, self._ro(f"{st['s']:.4f}"))
            self.table.setItem(r, 5, self._ro(f"{st['tp']:.3f}"))
            self.table.setItem(r, 6, self._ro(f"{st['tm']:.3f}"))
            self.table.setItem(r, 7, self._ro(f"{st['tn']:.3f}"))
            self.table.setItem(r, 8, self._ro(str(st["n"])))
            self.table.setItem(r, 9, self._ro(f"{st['area']:.2f}"))
        self.table.itemChanged.connect(self._label_edited)
        self._label_signal_connected = True

    def _label_edited(self, item):
        if item.column() == 2:
            r = item.row()
            if 0 <= r < len(self.cluster_stats):
                self.cluster_stats[r]["label"] = item.text()
                if self.mode == "cursor" and r < len(self.phasor.cursors):
                    self.phasor.cursors[r]["label"] = item.text()
                    self._refresh_active_cursor_combo(select_idx=r)

    @staticmethod
    def _ro(text):
        it = QtWidgets.QTableWidgetItem(text); it.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable); return it

    @staticmethod
    def _editable_group(text):
        it = QtWidgets.QTableWidgetItem(text)
        it.setFlags(
            Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEditable)
        it.setToolTip("Group name for this file (e.g. Tumor, Control)")
        return it

    @staticmethod
    def _rgb_hex(c):
        return "FF" + "".join(f"{int(round(255 * max(0.0, min(1.0, x)))):02X}" for x in c[:3])

    # ---- export ------------------------------------------------------------
    def save_overlay(self):
        if self.last_overlay is None: return
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Save overlay", "segmentation.png", "PNG (*.png)")
        if path:
            import matplotlib.pyplot as plt
            plt.imsave(path, self.last_overlay); self._log(f"Saved overlay: {path}")

    def save_phasor(self):
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Save phasor", "phasor.png", "PNG (*.png)")
        if path:
            self.phasor.fig.savefig(path, dpi=200); self._log(f"Saved phasor: {path}")

    def export_csv(self):
        if not self.cluster_stats: return
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Export results", "clusters.csv", "CSV (*.csv)")
        if not path: return
        with open(path, "w", newline="") as f:
            w = csv.writer(f)
            w.writerow(["cluster", "label", "g", "s", "tau_phi_ns", "tau_mod_ns", "tau_normal_ns",
                        "pixels", "area_percent", "frequency_MHz", "harmonic",
                        "sample_channel", "ref_channel", "filter", "group", "sample", "reference"])
            for st in self.cluster_stats:
                w.writerow([st["idx"], st["label"], f"{st['g']:.5f}", f"{st['s']:.5f}",
                            f"{st['tp']:.4f}", f"{st['tm']:.4f}", f"{st['tn']:.4f}",
                            st["n"], f"{st['area']:.3f}",
                            f"{self.data.work_frequency:.4f}", self.data.harmonic,
                            self.data.channel,
                            self.data.ref_channel if self.data.ref_path else "",
                            self.cb_filter.currentText(),
                            (self.data.group_name or "").strip(),
                            os.path.basename(self.data.sample_path),
                            os.path.basename(self._effective_ref_path() or "")])
        self._log(f"Exported {path}")

    def export_xlsx(self):
        if not self.cluster_stats:
            QtWidgets.QMessageBox.information(self, "Excel", "Nothing to export yet — compute clusters first.")
            return
        path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, "Export results (Excel)", "clusters.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            QtWidgets.QMessageBox.warning(
                self, "Missing dependency",
                "Excel export needs openpyxl:\n\n    pip install openpyxl")
            return
        try:
            import datetime
            wb = openpyxl.Workbook()
            ws = wb.active; ws.title = "Clusters"
            headers = ["#", "Color", "Label (what you see)", "g", "s",
                       "tau_phi (ns)", "tau_mod (ns)", "tau_normal (ns)",
                       "Pixels", "Area %"]
            ws.append(headers)
            for c in ws[1]:
                c.font = Font(bold=True); c.alignment = Alignment(horizontal="center")
            for st in self.cluster_stats:
                ws.append([st["idx"], "", st["label"],
                           round(st["g"], 5), round(st["s"], 5),
                           round(st["tp"], 4), round(st["tm"], 4), round(st["tn"], 4),
                           int(st["n"]), round(st["area"], 3)])
                ws.cell(row=ws.max_row, column=2).fill = PatternFill(
                    "solid", fgColor=self._rgb_hex(st["color"]))
            for i, w in enumerate([5, 8, 28, 10, 10, 12, 12, 14, 11, 9], start=1):
                ws.column_dimensions[get_column_letter(i)].width = w
            ws.freeze_panes = "A2"

            meta = wb.create_sheet("Metadata")
            meta.append(["Parameter", "Value"])
            for c in meta[1]:
                c.font = Font(bold=True)
            for k, v in [
                ("Sample", os.path.basename(self.data.sample_path)),
                ("Group", (self.data.group_name or "").strip() or "(none)"),
                ("Reference", os.path.basename(self.data.ref_path) if self.data.ref_path else ""),
                ("Laser frequency (MHz)", round(self.data.frequency, 4)),
                ("Working frequency (MHz)", round(self.data.work_frequency, 4)),
                ("Harmonic", self.data.harmonic),
                ("Sample channel", self.data.channel),
                ("Reference channel", self.data.ref_channel if self.data.ref_path else ""),
                ("Filter", self.cb_filter.currentText()),
                ("Segmentation mode", self.mode),
                ("Number of clusters", len(self.cluster_stats)),
                ("Exported", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ]:
                meta.append([k, v])
            meta.column_dimensions["A"].width = 26
            meta.column_dimensions["B"].width = 42
            wb.save(path)
            self._log(f"Exported {path}")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Excel export error", repr(e))

